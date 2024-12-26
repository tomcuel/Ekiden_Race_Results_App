# import the necessary libraries
import tkinter as tk
from tkinter import ttk
import os
import openpyxl
import re

# import the class created in other files
from utility import Utility


# Screen 2: classe to manage the sign up screen
class SignupScreen(tk.Frame):

    # Initialize the Sign Up Screen
    def __init__(self, master, navigate_callback):
        super().__init__(master, bg="#282c34", highlightthickness=0, borderwidth=0, border=0)
        self.navigate_callback = navigate_callback
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # Title Label
        label = tk.Label(
            self, text="Sign Up", font=("Helvetica", 70, "bold"), fg="#ffffff", bg="#282c34"
            )
        label.pack(pady=(20, 20))   

        # User Info Section
        user_info_frame = tk.LabelFrame(
            self, text="User Information", font=("Helvetica", 30), fg="#ffffff", bg="#282c34", border=0, labelanchor="n"
        )
        user_info_frame.pack(pady=(10,30), fill="y")

        first_name_label = tk.Label(
            user_info_frame, text="First Name (required)", font=("Helvetica", 20), fg="#ffffff", bg="#282c34"
        )
        first_name_label.grid(row=0, column=0)
        self.first_name_entry = tk.Entry(user_info_frame)
        self.first_name_entry.grid(row=1, column=0)
        self.first_name_entry.bind("<Return>", Utility.focus_next_widget)
        
        last_name_label = tk.Label(
            user_info_frame, text="Last Name (required)", font=("Helvetica", 20), fg="#ffffff", bg="#282c34"
        )
        last_name_label.grid(row=0, column=1)
        self.last_name_entry = tk.Entry(user_info_frame)
        self.last_name_entry.grid(row=1, column=1)
        self.last_name_entry.bind("<Return>", Utility.focus_next_widget)

        title_label = tk.Label(
            user_info_frame, text="Title", font=("Helvetica", 20), fg="#ffffff", bg="#282c34"
        )
        title_label.grid(row=0, column=2)
        self.title_combobox = ttk.Combobox(
            user_info_frame, values=["Mr.", "Ms.", "Dr."], state="readonly"
        )
        self.title_combobox.set('')
        self.title_combobox.grid(row=1, column=2)
        self.title_combobox.bind("<Return>", Utility.focus_next_widget)

        age_label = tk.Label(
            user_info_frame, text="Age", font=("Helvetica", 20), fg="#ffffff", bg="#282c34"
        )
        age_label.grid(row=2, column=0)
        self.age_combobox = ttk.Combobox(
            user_info_frame, values=[str(age) for age in range(15, 101)], state="readonly"
        )
        self.age_combobox.grid(row=3, column=0)
        self.age_combobox.bind("<Return>", Utility.focus_next_widget)

        nationality_label = tk.Label(
            user_info_frame, text="Nationality", font=("Helvetica", 20), fg="#ffffff", bg="#282c34"
            )
        nationality_label.grid(row=2, column=1)
        self.nationality_combobox = ttk.Combobox(
            user_info_frame, values=["French", "Tunisian", "Brazilian", "Ukrainian", "Polish", "Swiss", "Lebanese", "Other"], state="readonly"
            )
        self.nationality_combobox.set('')
        self.nationality_combobox.grid(row=3, column=1)
        self.nationality_combobox.bind("<Return>", Utility.focus_next_widget)

        for widget in user_info_frame.winfo_children():
            widget.grid_configure(padx=10, pady=5)

        # Runner Info Section
        runner_info_frame = tk.LabelFrame(
            self, text="Runner Information (required)", font=("Helvetica", 30), fg="#ffffff", bg="#282c34", border=0, labelanchor="n"
        )
        runner_info_frame.pack(pady=(10,30), fill="y")

        school_label = tk.Label(
            runner_info_frame, text="School", font=("Helvetica", 20), fg="#ffffff", bg="#282c34"
        )
        school_label.grid(row=0, column=0)
        self.school_combobox = ttk.Combobox(
            runner_info_frame, values=["Ensta Paris", "Télécom Paris", "ENSAE", "Polytechnique", "AgroParistech", "Centrale Supelec", "ENS Paris-Saclay", "Other"], state="readonly"
        )
        self.school_combobox.grid(row=1, column=0)
        self.school_combobox.bind("<Return>", Utility.focus_next_widget)

        self.runner_type_var = tk.StringVar(value="")
        runner_type_label = tk.Label(
            runner_info_frame, text="Type of Runner ", font=("Helvetica", 20), fg="#ffffff", bg="#282c34"
        )
        runner_type_label.grid(row=0, column=1, columnspan=2)
        student_button = tk.Radiobutton(
            runner_info_frame, text="Student", font=("Helvetica", 20), fg="#ffffff", bg="#282c34", variable=self.runner_type_var, value="Student"
        )
        student_button.grid(row=1, column=1)
        professor_button = tk.Radiobutton(
            runner_info_frame, text="Professor", font=("Helvetica", 20), fg="#ffffff", bg="#282c34", variable=self.runner_type_var, value="Professor"
        )
        professor_button.grid(row=1, column=2)
        student_button.bind("<Return>", Utility.focus_next_widget) # those can't be focused with the tab key
        professor_button.bind("<Return>", Utility.focus_next_widget) # those can't be modified by pressing a key

        for widget in runner_info_frame.winfo_children():
            widget.grid_configure(padx=10, pady=5)

        # Username and Password Info
        login_info_frame = tk.LabelFrame(
            self, text="Login Information (required)", font=("Helvetica", 30), fg="#ffffff", bg="#282c34", border=0, labelanchor="n"
        )
        login_info_frame.pack(pady=(10,30), fill="y")

        username_label = tk.Label(
            login_info_frame, text="Username", font=("Helvetica", 20), fg="#ffffff", bg="#282c34"
        )
        username_label.grid(row=0, column=0)
        self.username_entry = tk.Entry(login_info_frame)
        self.username_entry.grid(row=1, column=0)
        self.username_entry.bind("<Return>", Utility.focus_next_widget)

        password_label = tk.Label(
            login_info_frame, text="Password", font=("Helvetica", 20), fg="#ffffff", bg="#282c34"
        )
        password_label.grid(row=0, column=1)
        self.password_entry = tk.Entry(login_info_frame, show="*")
        self.password_entry.grid(row=1, column=1)
        self.password_entry.bind("<Return>", Utility.focus_next_widget)

        password_confirm_label = tk.Label(
            login_info_frame, text="Confirm Password", font=("Helvetica", 20), fg="#ffffff", bg="#282c34"
        )
        password_confirm_label.grid(row=0, column=2)
        self.password_confirm_entry = tk.Entry(login_info_frame, show="*")
        self.password_confirm_entry.grid(row=1, column=2)
        self.password_confirm_entry.bind("<Return>", Utility.focus_next_widget)

        for widget in login_info_frame.winfo_children():
            widget.grid_configure(padx=10, pady=5)

        # Accept Terms and Conditions
        terms_frame = tk.LabelFrame(
            self, text="Terms & Conditions (required)", font=("Helvetica", 20), fg="#ffffff", bg="#282c34"
        )
        terms_frame.pack(padx=20, pady=10, fill="x")
        self.accept_var = tk.StringVar(value="Not Accepted")
        terms_check = tk.Checkbutton(
            terms_frame, text="I accept the terms and conditions", font=("Helvetica", 10), fg="#ffffff", bg="#282c34",
            variable=self.accept_var, onvalue="Accepted", offvalue="Not Accepted"
        )
        terms_check.grid(row=0, column=0)
        terms_check.bind("<Return>", Utility.focus_next_widget)

        # Buttons
        self.submit_button = tk.Button(
            self, text="Submit", font=("Helvetica", 30), fg="#000000", relief="flat", activeforeground="#ff0000", border=0, highlightthickness=2.5, command=self.sign_up
        )
        self.submit_button.pack(pady=(20, 20))
        self.submit_button.bind("<Return>", lambda e: self.submit_button.invoke())

        self.back_button = tk.Button(
            self, text="Back to Login", font=("Helvetica", 30), fg="#000000", relief="flat", activeforeground="#ff0000", border=0, highlightthickness=2.5, command=lambda: self.navigate_callback(1),
        )
        self.back_button.pack(pady=(20, 20))
        self.back_button.bind("<Return>", lambda e: self.back_button.invoke())
        self.back_button.bind("<Tab>", lambda e: self.refocus_to_firstname())

    # Function to sign up a new user and save the data in the Excel file
    def sign_up(self):
        # Check acceptance of terms and conditions
        accepted = self.accept_var.get()
        if accepted == "Accepted":

            # User info
            firstname = self.first_name_entry.get()
            lastname = self.last_name_entry.get()
            # raising error if first name or last name is empty
            if not firstname:
                Utility.show_dismissable_messagebox(self, "Error", "First name is required", lambda: None)
                return
            if not lastname:
                Utility.show_dismissable_messagebox(self, "Error", "Last name is required", lambda: None)
                return
            
            # Runner info (title, age, nationality type of runner, school)
            title = self.title_combobox.get()
            age = self.age_combobox.get()
            nationality = self.nationality_combobox.get()
            # --> those are optional
            type_of_runner = self.runner_type_var.get()
            school = self.school_combobox.get()
            # raising error if type of runner, school, nationality is empty
            if type_of_runner == "":
                Utility.show_dismissable_messagebox(self, "Error", "Type of runner is required", lambda: None)
                return
            if school == "":
                Utility.show_dismissable_messagebox(self, "Error", "School is required", lambda: None)
                return

            # Username and password
            username = self.username_entry.get()
            password = self.password_entry.get()
            password_confirm = self.password_confirm_entry.get()
            # raising error if self.username or password is empty
            if not username or not password or not password_confirm:
                Utility.show_dismissable_messagebox(self, "Error", "Username and passwords are required", lambda: None)
                return
            # raising error if password and confirm password do not match
            if password != password_confirm:
                Utility.show_dismissable_messagebox(self, "Error", "Passwords do not match", lambda: None)
                return
            # the password should be at least 6 characters long, contain 2 numbers, and include at least one special character from *_+=?&@#-
            if not self.validate_password(password):
                Utility.show_dismissable_messagebox(self, "Error", "Password must be at least 6 characters long, contain 2 numbers, and include at least one special character from *_+=?&@#-", lambda: None)
                return

            # Filling the data in the excel file
            filepath = "Data/runner_data.xlsx"
            
            # Check if file exists, otherwise create a new one
            if not os.path.exists(filepath):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                heading = ["Title", "First Name", "Last Name", "Age", "Nationality", "Type of Runner", "School", "Username", "Password"]
                sheet.append(heading)
                workbook.save(filepath)

            # Load the workbook and sheet
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            
            # Check if the data already exists in the excel file (Full duplicate check)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if (row[1] == firstname and row[2] == lastname) or row[7] == username:
                    Utility.show_dismissable_messagebox(self, "Error", "This user already exists", lambda: None)
                    return
                if row[8] == password:  # Check for duplicate password
                    Utility.show_dismissable_messagebox(self, "Error", "This password is already used", lambda: None)
                    return

            # If no duplicates found, add new entry
            sheet.append([title, firstname, lastname, age, nationality, type_of_runner, school, username, password])
            workbook.save(filepath)

            # Reset the fields after submission
            self.first_name_entry.delete(0, 'end')
            self.last_name_entry.delete(0, 'end')
            self.title_combobox.set("")
            self.age_combobox.set('')
            self.nationality_combobox.set("")
            self.runner_type_var.set("")
            self.school_combobox.set("")
            self.username_entry.delete(0, 'end')
            self.password_entry.delete(0, 'end')
            self.password_confirm_entry.delete(0, 'end')
            self.accept_var.set("Not Accepted")

            Utility.show_dismissable_messagebox(self, "Success", "Sign-up successful", lambda: self.navigate_callback(1))
            self.navigate_callback(1)
        
        # If the terms are not accepted, show a warning
        else:
            Utility.show_dismissable_messagebox(self, "Error", "You have not accepted the terms", lambda: None)
    
      # Function to focus the next widget in the tab order when the Enter key is pressed
  
    # Function to validate the password
    def validate_password(self, password):
        return len(password) >= 8 and len(re.findall(r'[0-9]', password)) >= 2 and len(re.findall(r'[a-z]', password))+len(re.findall(r'[A-Z]', password)) >= 6 and any(char in "*_+=?&@#-" for char in password)

    # Function to refocus on the first name entry field when the Tab key is pressed on the Sign Up button
    def refocus_to_firstname(self):
        self.first_name_entry.focus()
        return "break"  
