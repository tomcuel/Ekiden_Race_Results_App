import tkinter  # for GUI
from tkinter import ttk  # for combobox
from tkinter import messagebox  # for error message
import os  # for file path
import openpyxl  # for excel file
import re  # for password validation

def run_data_entry_form():
    def validate_password(password):
        # Password must include at least 6 letters, 2 numbers, and 1 special character
        if len(password) < 6:
            return False
        if len(re.findall(r'[0-9]', password)) < 2:
            return False
        if not any(char in "*_+=?&@#-" for char in password):
            return False
        if not any(char.isalpha() for char in password):
            return False
        return True
    
    def enter_data():
        # Check acceptance of terms and conditions
        accepted = accept_var.get()
        if accepted == "Accepted":

            # User info
            firstname = first_name_entry.get()
            lastname = last_name_entry.get()
            # raising error if first name or last name is empty
            if not firstname:
                tkinter.messagebox.showwarning(title="Error", message="First name is required.")
                return
            if not lastname:
                tkinter.messagebox.showwarning(title="Error", message="Last name is required.")
                return
            
            # Runner info (title, age, nationality type of runner, school)
            title = title_combobox.get()
            age = age_combobox.get()
            nationality = nationality_combobox.get()
            # --> those are optional
            type_of_runner = runner_type_var.get()
            school = school_combobox.get()
            # raising error if type of runner, school, nationality is empty
            if type_of_runner == "":
                tkinter.messagebox.showwarning(title="Error", message="Type of runner is required.")
                return
            if school == "":
                tkinter.messagebox.showwarning(title="Error", message="School is required.")
                return

            # Username and password
            username = username_entry.get()
            password = password_entry.get()
            password_confirm = password_confirm_entry.get()
            # raising error if username or password is empty
            if not username or not password or not password_confirm:
                tkinter.messagebox.showwarning(title="Error", message="Username and passwords are required.")
                return
            # raising error if password and confirm password do not match
            if password != password_confirm:
                tkinter.messagebox.showwarning(title="Error", message="Passwords do not match.")
                return
            # the password should be at least 6 characters long, contain 2 numbers, and include at least one special character from *_+=?&@#-
            if not validate_password(password):
                tkinter.messagebox.showwarning(title="Error", message="Password must be at least 6 characters long, contain 2 numbers, and include at least one special character from *_+=?&@#-")
                return
            

            # Filling the data in the excel file
            filepath = "runner_data.xlsx"
            
            # Check if file exists, otherwise create a new one
            if not os.path.exists(filepath):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                heading = ["Title", "First Name", "Last Name", "Age", "Nationality", "Type of Runner", "School", "Username", "Password"]
                sheet.append(heading)
                workbook.save(filepath)

            # Load the workbook and sheet
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            
            # Check if the data already exists in the excel file (Full duplicate check)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if (row[1] == firstname and row[2] == lastname) or row[7] == username:
                    tkinter.messagebox.showwarning(title="Error", message="This user already exists.")
                    return
                if row[8] == password:  # Check for duplicate password
                    tkinter.messagebox.showwarning(title="Error", message="This password is already used.")
                    return

            # If no duplicates found, add new entry
            sheet.append([title, firstname, lastname, age, nationality, type_of_runner, school, username, password])
            workbook.save(filepath)

            # Reset the fields after submission
            first_name_entry.delete(0, 'end')
            last_name_entry.delete(0, 'end')
            title_combobox.set("")
            age_combobox.set('')
            nationality_combobox.set("")
            runner_type_var.set("")
            school_combobox.set("")
            username_entry.delete(0, 'end')
            password_entry.delete(0, 'end')
            password_confirm_entry.delete(0, 'end')
            accept_var.set("Not Accepted")

            tkinter.messagebox.showinfo(title="Success", message="Data entered successfully!")
        else:
            tkinter.messagebox.showwarning(title="Error", message="You have not accepted the terms.")
    
    window = tkinter.Tk()
    window.title("Data Entry Form")

    frame = tkinter.Frame(window)
    frame.pack()

    # User Info
    user_info_frame = tkinter.LabelFrame(frame, text="User Information")
    user_info_frame.grid(row=0, column=0, padx=20, pady=10)

    first_name_label = tkinter.Label(user_info_frame, text="First Name (required)")
    first_name_label.grid(row=0, column=0)
    last_name_label = tkinter.Label(user_info_frame, text="Last Name (required)")
    last_name_label.grid(row=0, column=1)

    first_name_entry = tkinter.Entry(user_info_frame)
    last_name_entry = tkinter.Entry(user_info_frame)
    first_name_entry.grid(row=1, column=0)
    last_name_entry.grid(row=1, column=1)

    title_label = tkinter.Label(user_info_frame, text="Title")
    title_combobox = ttk.Combobox(user_info_frame, values=["Mr.", "Ms.", "Dr."])
    title_combobox.set('')  # Make it so no default is pre-selected.
    title_combobox['state'] = 'readonly'  # Set to readonly to prevent typing.
    title_label.grid(row=0, column=2)
    title_combobox.grid(row=1, column=2)

    # Age Combobox with restricted values from 15 to 100
    age_label = tkinter.Label(user_info_frame, text="Age")
    age_combobox = ttk.Combobox(user_info_frame, values=[str(age) for age in range(15, 101)], state="readonly")
    age_label.grid(row=2, column=0)
    age_combobox.grid(row=3, column=0)

    nationality_label = tkinter.Label(user_info_frame, text="Nationality")
    nationality_combobox = ttk.Combobox(user_info_frame, 
                                         values=["French", "Tunisian", "Brazilian", "Ukrainian", "Polish", 
                                                 "Swiss", "Lebanese", "Other"])
    nationality_combobox.set('')  # Make it so no default is pre-selected.
    nationality_combobox['state'] = 'readonly'  # Set to readonly to prevent typing.
    nationality_label.grid(row=2, column=1)
    nationality_combobox.grid(row=3, column=1)

    for widget in user_info_frame.winfo_children():
        widget.grid_configure(padx=10, pady=5)

    # Runner Info
    runner_info_frame = tkinter.LabelFrame(frame, text="Runner Information")
    runner_info_frame.grid(row=1, column=0, sticky="news", padx=20, pady=10)
    runner_type_label = tkinter.Label(runner_info_frame, text="Type of Runner (required)")
    runner_type_var = tkinter.StringVar(value="")
    student_button = tkinter.Radiobutton(runner_info_frame, text="Student", variable=runner_type_var, value="Student")
    professor_button = tkinter.Radiobutton(runner_info_frame, text="Professor", variable=runner_type_var, value="Professor")
    runner_type_label.grid(row=0, column=0)
    
    student_button.grid(row=1, column=0)
    professor_button.grid(row=1, column=1)
    
    school_label = tkinter.Label(runner_info_frame, text="School (required)")
    school_combobox = ttk.Combobox(runner_info_frame, values=["Ensta Paris", "Télécom Paris", "ENSAE", "Polytechnique", "AgroParistech", "Centrale Supelec", "ENS Paris-Saclay", "Other"], state="readonly")
    school_label.grid(row=0, column=2)
    school_combobox.grid(row=1, column=2)
    
    for widget in runner_info_frame.winfo_children():
        widget.grid_configure(padx=10, pady=5)

    # Username and Password Info
    login_info_frame = tkinter.LabelFrame(frame, text="Login Information")
    login_info_frame.grid(row=2, column=0, sticky="news", padx=20, pady=10)

    username_label = tkinter.Label(login_info_frame, text="Username (required)")
    username_label.grid(row=0, column=0)
    username_entry = tkinter.Entry(login_info_frame)
    username_entry.grid(row=1, column=0)

    password_label = tkinter.Label(login_info_frame, text="Password (required)")
    password_label.grid(row=0, column=1)
    password_entry = tkinter.Entry(login_info_frame, show="*")
    password_entry.grid(row=1, column=1)

    password_confirm_label = tkinter.Label(login_info_frame, text="Confirm Password (required)")
    password_confirm_label.grid(row=0, column=2)
    password_confirm_entry = tkinter.Entry(login_info_frame, show="*")
    password_confirm_entry.grid(row=1, column=2)

    for widget in login_info_frame.winfo_children():
        widget.grid_configure(padx=10, pady=5)

    # Accept Terms and Conditions
    terms_frame = tkinter.LabelFrame(frame, text="Terms & Conditions (required)")
    terms_frame.grid(row=3, column=0, sticky="news", padx=20, pady=10)

    accept_var = tkinter.StringVar(value="Not Accepted")
    terms_check = tkinter.Checkbutton(terms_frame, text="I accept the terms and conditions.",
                                      variable=accept_var, onvalue="Accepted", offvalue="Not Accepted")
    terms_check.grid(row=0, column=0)

    # Submit Button
    button = tkinter.Button(frame, text="Enter Data", command=enter_data)
    button.grid(row=4, column=0, sticky="news", padx=20, pady=10)

    window.mainloop()

# Call the function to run the form
run_data_entry_form()
