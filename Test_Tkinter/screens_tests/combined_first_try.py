import tkinter as tk
from tkinter import ttk, messagebox
import os
import openpyxl
import re


# Screen 1: Login Screen
class LoginScreen(tk.Frame):
    # Initialize the Login Screen
    def __init__(self, master, navigate_callback):
        super().__init__(master)
        self.navigate_callback = navigate_callback

        self.label = tk.Label(self, text="Login", font=("Helvetica", 24))
        self.label.pack(pady=20)

        self.username_label = tk.Label(self, text="Username:")
        self.username_label.pack()
        self.username_entry = tk.Entry(self)
        self.username_entry.pack()

        self.password_label = tk.Label(self, text="Password:")
        self.password_label.pack()
        self.password_entry = tk.Entry(self, show="*")
        self.password_entry.pack()

        self.login_button = tk.Button(self, text="Login", command=self.login)
        self.login_button.pack(pady=10)

        self.signup_button = tk.Button(self, text="Sign Up", command=lambda: self.navigate_callback(2))
        self.signup_button.pack()

    # Function to check if the login credentials are valid (in the Excel file)
    def login(self):
        username = self.username_entry.get()
        password = self.password_entry.get()

        # Open the Excel file and get the active sheet
        try:
            workbook = openpyxl.load_workbook('runner_data.xlsx')  # Adjust the path to your actual Excel file
            sheet = workbook.active
            
            # Loop through each row in the sheet (starting from row 2 to skip the header row)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Check if the input username and password match the stored values
                if row[7] == username and row[8] == password:
                    messagebox.showinfo("Success", "Login successful!")
                    self.navigate_callback(3)
                    return
            # If no match is found, show an error
            messagebox.showerror(title="Error", message="Invalid login.")
        
        except Exception as e:
            messagebox.showerror(title="Error", message=f"An error occurred: {e}")


class SignupScreen(tk.Frame):
    # Initialize the Sign Up Screen
    def __init__(self, master, navigate_callback):
        super().__init__(master)
        self.navigate_callback = navigate_callback

        # Title Label
        label = tk.Label(self, text="Sign Up Form", font=("Helvetica", 24))
        label.pack(pady=20)

        # User Info Section
        user_info_frame = tk.LabelFrame(self, text="User Information")
        user_info_frame.pack(padx=20, pady=10, fill="x")

        first_name_label = tk.Label(user_info_frame, text="First Name (required)")
        first_name_label.grid(row=0, column=0)
        last_name_label = tk.Label(user_info_frame, text="Last Name (required)")
        last_name_label.grid(row=0, column=1)

        self.first_name_entry = tk.Entry(user_info_frame)
        self.first_name_entry.grid(row=1, column=0)
        self.last_name_entry = tk.Entry(user_info_frame)
        self.last_name_entry.grid(row=1, column=1)

        title_label = tk.Label(user_info_frame, text="Title")
        title_label.grid(row=0, column=2)
        self.title_combobox = ttk.Combobox(user_info_frame, values=["Mr.", "Ms.", "Dr."], state="readonly")
        self.title_combobox.set('')
        self.title_combobox.grid(row=1, column=2)

        age_label = tk.Label(user_info_frame, text="Age")
        age_label.grid(row=2, column=0)
        self.age_combobox = ttk.Combobox(user_info_frame, values=[str(age) for age in range(15, 101)], state="readonly")
        self.age_combobox.grid(row=3, column=0)

        nationality_label = tk.Label(user_info_frame, text="Nationality")
        nationality_label.grid(row=2, column=1)
        self.nationality_combobox = ttk.Combobox(user_info_frame, values=["French", "Tunisian", "Brazilian", "Ukrainian", "Polish", "Swiss", "Lebanese", "Other"], state="readonly")
        self.nationality_combobox.set('')
        self.nationality_combobox.grid(row=3, column=1)

        for widget in user_info_frame.winfo_children():
            widget.grid_configure(padx=10, pady=5)

        # Runner Info Section
        runner_info_frame = tk.LabelFrame(self, text="Runner Information")
        runner_info_frame.pack(padx=20, pady=10, fill="x")

        self.runner_type_var = tk.StringVar(value="")
        runner_type_label = tk.Label(runner_info_frame, text="Type of Runner (required)")
        runner_type_label.grid(row=0, column=0)

        student_button = tk.Radiobutton(runner_info_frame, text="Student", variable=self.runner_type_var, value="Student")
        professor_button = tk.Radiobutton(runner_info_frame, text="Professor", variable=self.runner_type_var, value="Professor")
        student_button.grid(row=1, column=0)
        professor_button.grid(row=1, column=1)

        school_label = tk.Label(runner_info_frame, text="School (required)")
        school_label.grid(row=0, column=2)
        self.school_combobox = ttk.Combobox(runner_info_frame, values=["Ensta Paris", "Télécom Paris", "ENSAE", "Polytechnique", "AgroParistech", "Centrale Supelec", "ENS Paris-Saclay", "Other"], state="readonly")
        self.school_combobox.grid(row=1, column=2)

        for widget in runner_info_frame.winfo_children():
            widget.grid_configure(padx=10, pady=5)

        # Username and Password Info
        self.login_info_frame = tk.LabelFrame(self, text="Login Information")
        self.login_info_frame.pack(padx=20, pady=10, fill="x")

        username_label = tk.Label(self.login_info_frame, text="Username (required)")
        username_label.grid(row=0, column=0)
        self.username_entry = tk.Entry(self.login_info_frame)
        self.username_entry.grid(row=1, column=0)

        password_label = tk.Label(self.login_info_frame, text="Password (required)")
        password_label.grid(row=0, column=1)
        self.password_entry = tk.Entry(self.login_info_frame, show="*")
        self.password_entry.grid(row=1, column=1)

        password_confirm_label = tk.Label(self.login_info_frame, text="Confirm Password (required)")
        password_confirm_label.grid(row=0, column=2)
        self.password_confirm_entry = tk.Entry(self.login_info_frame, show="*")
        self.password_confirm_entry.grid(row=1, column=2)

        for widget in self.login_info_frame.winfo_children():
            widget.grid_configure(padx=10, pady=5)

        # Accept Terms and Conditions
        self.terms_frame = tk.LabelFrame(self, text="Terms & Conditions (required)")
        self.terms_frame.pack(padx=20, pady=10, fill="x")

        self.accept_var = tk.StringVar(value="Not Accepted")
        terms_check = tk.Checkbutton(self.terms_frame, text="I accept the terms and conditions.",
                                     variable=self.accept_var, onvalue="Accepted", offvalue="Not Accepted")
        terms_check.grid(row=0, column=0)

        # Buttons
        self.submit_button = tk.Button(self, text="Submit", command=self.sign_up)
        self.submit_button.pack(pady=10)

        self.back_button = tk.Button(self, text="Back to Login", command=lambda: self.navigate_callback(1))
        self.back_button.pack(pady=10)

    # Function to sign up a new user and save the data in the Excel file
    def sign_up(self):
         # Check acceptance of terms and conditions
        accepted = self.accept_var.get()
        if accepted == "Accepted":

            # User info
            firstname = self.first_name_entry.get()
            lastname = self.last_name_entry.get()
            # raising error if first name or last name is empty
            if not firstname:
                tk.messagebox.showwarning(title="Error", message="First name is required.")
                return
            if not lastname:
                tk.messagebox.showwarning(title="Error", message="Last name is required.")
                return
            
            # Runner info (title, age, nationality type of runner, school)
            title = self.title_combobox.get()
            age = self.age_combobox.get()
            nationality = self.nationality_combobox.get()
            # --> those are optional
            type_of_runner = self.runner_type_var.get()
            school = self.school_combobox.get()
            # raising error if type of runner, school, nationality is empty
            if type_of_runner == "":
                tk.messagebox.showwarning(title="Error", message="Type of runner is required.")
                return
            if school == "":
                tk.messagebox.showwarning(title="Error", message="School is required.")
                return

            # Username and password
            username = self.username_entry.get()
            password = self.password_entry.get()
            password_confirm = self.password_confirm_entry.get()
            # raising error if self.username or password is empty
            if not username or not password or not password_confirm:
                tk.messagebox.showwarning(title="Error", message="Username and passwords are required.")
                return
            # raising error if password and confirm password do not match
            if password != password_confirm:
                tk.messagebox.showwarning(title="Error", message="Passwords do not match.")
                return
            # the password should be at least 6 characters long, contain 2 numbers, and include at least one special character from *_+=?&@#-
            if not self.validate_password(password):
                tk.messagebox.showwarning(title="Error", message="Password must be at least 6 characters long, contain 2 numbers, and include at least one special character from *_+=?&@#-")
                return

            # Filling the data in the excel file
            filepath = "runner_data.xlsx"
            
            # Check if file exists, otherwise create a new one
            if not os.path.exists(filepath):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                heading = ["Title", "First Name", "Last Name", "Age", "Nationality", "Type of Runner", "School", "Username", "Password"]
                sheet.append(heading)
                workbook.save(filepath)

            # Load the workbook and sheet
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            
            # Check if the data already exists in the excel file (Full duplicate check)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if (row[1] == firstname and row[2] == lastname) or row[7] == username:
                    tk.messagebox.showwarning(title="Error", message="This user already exists.")
                    return
                if row[8] == password:  # Check for duplicate password
                    tk.messagebox.showwarning(title="Error", message="This password is already used.")
                    return

            # If no duplicates found, add new entry
            sheet.append([title, firstname, lastname, age, nationality, type_of_runner, school, username, password])
            workbook.save(filepath)

            # Reset the fields after submission
            self.first_name_entry.delete(0, 'end')
            self.last_name_entry.delete(0, 'end')
            self.title_combobox.set("")
            self.age_combobox.set('')
            self.nationality_combobox.set("")
            self.runner_type_var.set("")
            self.school_combobox.set("")
            self.username_entry.delete(0, 'end')
            self.password_entry.delete(0, 'end')
            self.password_confirm_entry.delete(0, 'end')
            self.accept_var.set("Not Accepted")

            messagebox.showinfo("Success", "Sign-up successful!")
            self.navigate_callback(1)
        
        # If the terms are not accepted, show a warning
        else:
            tk.messagebox.showwarning(title="Error", message="You have not accepted the terms.")
       
    # Function to validate the password
    def validate_password(self, password):
        return len(password) >= 8 and len(re.findall(r'[0-9]', password)) >= 2 and len(re.findall(r'[a-z]', password))+len(re.findall(r'[A-Z]', password)) >= 6 and any(char in "*_+=?&@#-" for char in password)



# Screen 3: Blank Screen
class BlankScreen(tk.Frame):
    def __init__(self, master, navigate_callback):
        super().__init__(master)
        self.navigate_callback = navigate_callback

        self.label = tk.Label(self, text="Welcome to Screen 3!", font=("Helvetica", 24))
        self.label.pack(pady=20)


# Main Scrollable App
class ScrollableApp:
    def __init__(self, window):
        self.window = window
        self.window.title("Multi-Screen Scrollable App")

        # Set the window to full screen
        self.window.attributes("-fullscreen", True)

        # Create a container frame to hold all screens
        self.container = tk.Frame(window)
        self.container.pack(fill="both", expand=True)

        # Create and store screens in the container
        self.screens = {
            1: LoginScreen(self.container, self.show_screen),
            2: SignupScreen(self.container, self.show_screen),
            3: BlankScreen(self.container, self.show_screen)
        }

        # Place all screens in the same location within the container
        for screen in self.screens.values():
            screen.grid(row=0, column=0, sticky="nsew")

        # Start with the first screen
        self.current_screen = 1
        self.show_screen(self.current_screen)

        # Bind keyboard inputs
        self.window.bind("<Escape>", self.quit_game)

    def show_screen(self, screen_number):
        # Raise the selected screen to the front and update the current screen tracker
        self.screens[screen_number].tkraise()
        self.current_screen = screen_number

    def next_screen(self, event=None):
        if self.current_screen < len(self.screens):
            self.show_screen(self.current_screen + 1)

    def previous_screen(self, event=None):
        if self.current_screen > 1:
            self.show_screen(self.current_screen - 1)

    def quit_game(self, event=None):
        # Quit the game and restore the window to a normal state
        self.window.attributes("-fullscreen", False)
        self.window.quit()


# Create the main window
window = tk.Tk()
app = ScrollableApp(window)
window.mainloop()
