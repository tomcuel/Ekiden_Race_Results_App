from enum import Enum
import tkinter as tk
from tkinter import ttk, messagebox
from tkinter import PhotoImage, Scrollbar, Canvas
import os
import openpyxl
import re



# Define the teams_data class to parse the data and store the team information
class Teams:


    # Enum to define the team categories
    class team_type(Enum):
        MEN = 0
        WOMEN = 1
        MIXED = 2 # Admin is included in the mixed category
        UNKOWN = 4


    # Enum to define the fields of the team_info object
    class field(Enum):
        TEAM_NAME = 0
        TEAM_CATEGORY = 1
        TEAM_TIME = 2
        TEAM_RANKING_IN_CATEGORY = 3
        TEAM_OVERALL_RANKING = 4


    # Define the runner class to store the data of each runner
    class runner:
        def __init__(self, name: str, sex: int, passage_number: int, distance: float, medley_time: int, speed: float):
            self.name = name
            self.sex = Teams.team_type(sex)
            self.passage_number = passage_number
            self.distance = distance
            self.medley_time = medley_time  # Time in seconds
            self.speed = speed

        def __repr__(self):
            return (f"runner(name='{self.name}', sex='{self.sex.name}', "
                    f"passage_number={self.passage_number}, distance={self.distance}, "
                    f"medley_time={self.medley_time}, speed={self.speed})")


    # Define the team_info object to store the data of each team
    class team_info:
        # Initialize the team_info object with the team's name, category, time, runners, and ranking
        def __init__(self, name: str, category: int, time: int, category_ranking: int, ranking: int, runners: list):
            self.data = {
                Teams.field.TEAM_NAME: name,
                Teams.field.TEAM_CATEGORY: category,
                Teams.field.TEAM_TIME: time,
                Teams.field.TEAM_RANKING_IN_CATEGORY: category_ranking,
                Teams.field.TEAM_OVERALL_RANKING: ranking
            }
            self.runners = runners 

        # Define the string representation of the object
        def __repr__(self):
            category_name = Teams.team_type(self.data[Teams.field.TEAM_CATEGORY]).name
            category_ranking_display = self.data[Teams.field.TEAM_RANKING_IN_CATEGORY] if self.data[Teams.field.TEAM_RANKING_IN_CATEGORY] >= 0 else "Unranked"
            ranking_display = self.data[Teams.field.TEAM_OVERALL_RANKING] if self.data[Teams.field.TEAM_OVERALL_RANKING] >= 0 else "Unranked"
            return (f"team_info(team_name='{self.data[Teams.field.TEAM_NAME]}', "
                    f"team_category='{category_name}', "
                    f"team_time={self.data[Teams.field.TEAM_TIME]}, "
                    f"team_ranking_in_category={category_ranking_display}, "
                    f"team_overall_ranking={ranking_display}, "
                    f"runners={self.runners})")


    # Initialize the teams_data object with the file path
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.teams = self._parse_teams()

    # Define the string representation of the object
    def __repr__(self):
        return (f"teams_data(file_path='{self.file_path}', "
                f" team_data={self.teams})")

    # Parse the teams data from the file, return a list of team_info objects
    def _parse_teams(self):
        with open(self.file_path, "r") as file:
            data = file.read().strip().split("\n")

        # Remove the first 5 lines of unnecessary data
        data = data[5:]
        number_of_teams = data.count("") + 1

        # Initialize lists to store each team's data
        teams_data_parsing = [[] for _ in range(number_of_teams)]
        team_index = 0
        while data:
            line = data.pop(0)
            if line == "":
                team_index += 1
                continue
            teams_data_parsing[team_index].append(line)

        # Process each team into team_info objects
        teams = []
        for i, team_info_parsing in enumerate(teams_data_parsing):
            if not team_info_parsing:
                continue
            
            # The first line gives us the data of the team
            team_info_line = team_info_parsing[0].strip().split("|")
            
            # Extract the team name
            team_name = team_info_line[0]
            
            # Get the ekiden total time in seconds
            team_time_str = team_info_line[1].strip().split(" ")[3].split(":")
            team_time = int(team_time_str[0])*3600 + int(team_time_str[1])*60 + int(team_time_str[2])
            
            # Get the team category number
            # Non classed teams have an extra word in the category
            delta = 0
            if team_info_line[2].strip().split(" ")[0] == "Non" :
                delta = 1
            team_category_str = team_info_line[2].strip().split(" ")[1+delta]
            if team_category_str == "Hommes":
                category = Teams.team_type.MEN
            elif team_category_str == "Dames":
                category = Teams.team_type.WOMEN
            elif team_category_str == "Mixte":
                category = Teams.team_type.MIXED
            elif team_category_str == "Admin":
                category = Teams.team_type.MIXED
            else:
                category = Teams.team_type.UNKOWN
            
            # Get the team ranking in its category
            category_team_ranking = team_info_line[2].strip().split(" ")[0][:2]
            if category_team_ranking[1] == '°':
                category_team_ranking = category_team_ranking[0]
            if category_team_ranking == "No":
                category_team_ranking = -1
            else :
                category_team_ranking = int(category_team_ranking)
            
            # Get the team overall ranking
            team_ranking = i+1

            runners = []
            # Process the runners data, 6 runners for line 3/4/5/6/7/8
            for i in range(2, 8) : 
                
                runner_info = team_info_parsing[i].split("\t")

                # Name and Sex
                runner_name = runner_info[1].strip()
                runner_sex = Teams.team_type.MEN if runner_info[2] == "M" else Teams.team_type.WOMEN

                # Runner passage number
                runner_passage_number = i-1

                 # Runner distance
                runner_distance = 7.2
                if (runner_passage_number == 1 or runner_passage_number == 3 or runner_passage_number == 5):
                    runner_distance = 5.0
                elif (runner_passage_number == 2 or runner_passage_number == 4):
                    runner_distance = 10.0

                # Medley time 
                medley_time_str = runner_info[4]
                medley_time = 0
                # Check if the time includes hours (separating the hours, minutes, and seconds differently if there is hours or not)
                if not 'X' in medley_time_str:
                    if 'h' in medley_time_str:
                        medley_time = 3600 * int(medley_time_str[0]) + 60 * int(medley_time_str[2:4]) + int(medley_time_str[5:7])
                    else:
                        medley_time = 60 * int(medley_time_str[0:2]) + int(medley_time_str[3:5])
            
                # Speed
                runner_speed_str = runner_info[5].strip()
                if runner_speed_str and runner_speed_str.replace(",", ".").replace(".", "", 1).isdigit():
                    runner_speed = float(runner_speed_str.replace(",", "."))
                else:
                    runner_speed = 0.0  # Or handle as needed (e.g., raise an error or set a default value)

                # Append the runner to the runners list
                partial_runner = Teams.runner(name=runner_name, sex=runner_sex, passage_number=runner_passage_number, distance=runner_distance, medley_time=medley_time, speed=runner_speed)
                runners.append(partial_runner)


            # Create a team_info instance and add it to the teams list
            team = Teams.team_info(name=team_name, category=category, time=team_time, category_ranking=category_team_ranking, ranking=team_ranking, runners=runners)
            teams.append(team)

        # Return the list of team_info objects
        return teams

    # Function to get the teams
    def get_teams(self):
        return self.teams
    
    # Function to get the team based on its name (or at least a part of it)
    def get_team(self, team_name: str):
        for team in self.teams:
            if team_name in team.data[Teams.field.TEAM_NAME]:
                return team
        return None
    
    # Function to get the teams of one category
    def get_teams_by_category(self, category):
        return [team for team in self.teams if team.data[Teams.field.TEAM_CATEGORY] == category]
    
    # Function to get the teams of a category depending on the team name
    def get_teams_by_category_by_name(self, name : str):
        category_searched = -1
        for team in self.teams : 
            if name in team.data[Teams.field.TEAM_NAME] :
                category_searched = team.data[Teams.field.TEAM_CATEGORY]
                break
        return self.get_teams_by_category(category_searched)
    
    # Function to get the runners of a team
    def get_runners(self, team):
        return team.runners
    

    # Define the number relay class to store the data of each relay
    class Number_Relay:

        # Initialize the number relay object with the number of the relay and teams data
        def __init__(self, name : str, number: int, sex: int, teams: list):
            self.name = name
            self.number = number
            self.sex = sex
            self.teams = teams
            self.number_relay = self.get_number_relay(name, number, sex)

        # Define the string representation of the object
        def __repr__(self):
            return (f"Number_Relay(name='{self.name}', "
                    f"number='{self.number}', "
                    f"sex='{self.sex}', "
                    f"number_relay={self.number_relay})")

        # function to get the distance relay
        # the sex matter, because it will get the scratch time of the runners, or the ordering based on the sex
        def get_number_relay(self, name : str, number : int,  sex : int):
            self.name = name    
            self.number = number
            self.sex = sex
            relay_number = []
            # empty name means we want to get the relay based on the number, still looking at the sex
            if name == "":
                for team in self.teams:
                    if team.runners[self.number - 1].sex == sex or sex == Teams.team_type.MIXED : 
                        relay_number.append(team.runners[self.number - 1])
            # otherwise we want to get the relay based on the name, still looking at the sex
            else : 
                number_searched = -1
                sex_searched = -1
                for team in self.teams : 
                    for runner in team.runners : 
                        if name in runner.name : 
                            number_searched = runner.passage_number
                            self.number = number_searched
                            sex_searched = runner.sex
                            self.sex = sex_searched
                            break
                # if the number is not found, return an empty list since there is no runner with this name
                if number_searched == -1 :
                    return []
                for team in self.teams:
                    if team.runners[number_searched - 1].sex == sex or sex == Teams.team_type.MIXED : 
                        relay_number.append(team.runners[number_searched - 1])
            # return the list of the runners
            return relay_number


    # Define the distance relay class to store the data of each relay
    class Distance_Relay:

        # Initialize the distance relay object with the distance and teams data
        def __init__(self, name : str, distance: float, sex: int, teams: list):
            self.name = name
            self.distance = distance
            self.sex = sex
            self.teams = teams
            self.distance_relay = self.get_distance_relay(name, distance, sex)

        # Define the string representation of the object
        def __repr__(self):
            return (f"Distance_Relay(name='{self.name}', "
                    f"distance='{self.distance}', "
                    f"sex='{self.sex}', "
                    f"distance_relay={self.distance_relay})")

        # function to get the distance relay
        # the sex matter, because it will get the scratch time of the runners, or the ordering based on the sex
        # no name means that we want to get the relay based on the distance, still looking at the sex
        # if we want a WOMEN but the name correspond to a MEN, it will return the MIXED relay
        def get_distance_relay(self, name : str, distance : float, sex : int):
            self.name = name
            self.distance = distance
            self.sex = sex
            relay_distance = []
            # empty name means we want to get the relay based on the distance, still looking at the sex
            if name == "":
                for team in self.teams:
                    for runner in team.runners : 
                        if runner.distance == distance and (runner.sex == sex or sex == Teams.team_type.MIXED) : 
                            relay_distance.append(runner)
            # otherwise we want to get the relay based on the name, still looking at the sex
            else : 
                distance_searched = -1
                sex_searched = -1
                for team in self.teams : 
                    for runner in team.runners : 
                        if name in runner.name :
                            distance_searched = runner.distance
                            self.distance = distance_searched
                            sex_searched = runner.sex
                            self.sex = sex_searched
                            break
                # if the distance is not found, return an empty list since there is no runner with this name
                if distance_searched == -1 or sex_searched == -1 :
                    return []
                for team in self.teams:
                    for runner in team.runners : 
                        if runner.distance == distance_searched and (runner.sex == sex_searched or sex == Teams.team_type.MIXED) : 
                            relay_distance.append(runner)
            # return the list of the runners
            return relay_distance



# Utility class to gather some useful functions that are common to multiple screens
class Utility:
    @staticmethod
    # Function to show a dismissable messagebox with a custom duration, that can navigate to a new screen also, depending on the callback
    def show_dismissable_messagebox(parent, title, message, navigate_callback, duration = 3000):
        current_focus = parent.focus_get()

        # Create a top-level window
        popup = tk.Toplevel(parent)
        popup.title(title)
        popup.geometry("300x150")  
        popup.transient(parent)  
        popup.resizable(False, False)  
        popup.grab_set()

        # Add the message to the popup
        label = tk.Label(popup, text=message, font=("Helvetica", 12), wraplength=250, pady=20)
        label.pack()

        # Add a dismiss button
        dismiss_button = tk.Button(popup, text="Dismiss", command=lambda: (popup.destroy(), navigate_callback()))
        dismiss_button.pack(pady=10)

        # Automatically close the popup after `duration` milliseconds and navigate
        def auto_close():
            if popup.winfo_exists():  # Check if the popup is still open
                popup.destroy()
                navigate_callback()

        popup.after(duration, auto_close)

        # Return focus to the previously focused widget when the popup closes
        def on_close():
            if current_focus:
                current_focus.focus_set()

        popup.protocol("WM_DELETE_WINDOW", lambda: (on_close(), popup.destroy()))  # Handle the close button (X)
        popup.bind("<Destroy>", lambda event: on_close())
    
    @staticmethod
    # Function to focus the next widget in the tab order when the Enter key is pressed
    def focus_next_widget(event):
        # Focus the next widget in tab order
        event.widget.tk_focusNext().focus()
        return "break"



# Screen 1: Login Screen
class LoginScreen(tk.Frame):
    # Initialize the Login Screen
    def __init__(self, master, navigate_callback):
        super().__init__(master)
        self.navigate_callback = navigate_callback

        label = tk.Label(self, text="Login", font=("Helvetica", 24))
        label.pack()

        # Creating the screen fields
        username_label = tk.Label(self, text="Username:")
        username_label.pack()
        self.username_entry = tk.Entry(self)
        self.username_entry.pack()
        self.username_entry.bind("<Return>", Utility.focus_next_widget)

        password_label = tk.Label(self, text="Password:")
        password_label.pack()
        self.password_entry = tk.Entry(self, show="*")
        self.password_entry.pack()
        self.password_entry.bind("<Return>", Utility.focus_next_widget)

        self.login_button = tk.Button(self, text="Login", command=self.login)
        self.login_button.pack()
        self.login_button.bind("<Return>", lambda e: self.login_button.invoke())

        self.signup_button = tk.Button(self, text="Sign Up", command=lambda: self.navigate_callback(2))
        self.signup_button.pack()
        self.signup_button.bind("<Return>", lambda e: self.signup_button.invoke())
        self.signup_button.bind("<Tab>", lambda e: self.refocus_to_username())

    # Function to check if the login credentials are valid (in the Excel file)
    def login(self):
        username = self.username_entry.get()
        password = self.password_entry.get()

        # Open the Excel file and get the active sheet
        try:
            workbook = openpyxl.load_workbook('runner_data.xlsx')  # Adjust the path to your actual Excel file
            sheet = workbook.active
            
            # Loop through each row in the sheet (starting from row 2 to skip the header row)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Check if the input username and password match the stored values
                if row[7] == username and row[8] == password:
                    Utility.show_dismissable_messagebox(self, "Success", "Login successful!", lambda: self.navigate_callback(3))
                    return
            # If no match is found, show an error and reset the password fields
            Utility.show_dismissable_messagebox(self, "Error", "Invalid login", lambda: None)
            self.username_entry.delete(0, 'end')
            self.password_entry.delete(0, 'end')
        
        except Exception as e:
            messagebox.showerror(title="Error", message=f"An error occurred: {e}")

    # Function to refocus on the username entry field when the Tab key is pressed on the Sign Up button
    def refocus_to_username(self):
        self.username_entry.focus()
        return "break"  



# Screen 2: Sign Up Screen
class SignupScreen(tk.Frame):
    # Initialize the Sign Up Screen
    def __init__(self, master, navigate_callback):
        super().__init__(master)
        self.navigate_callback = navigate_callback

        # Title Label
        label = tk.Label(self, text="Sign Up Form", font=("Helvetica", 24))
        label.pack(pady=20)

        # User Info Section
        user_info_frame = tk.LabelFrame(self, text="User Information")
        user_info_frame.pack(padx=20, pady=10, fill="x")

        first_name_label = tk.Label(user_info_frame, text="First Name (required)")
        first_name_label.grid(row=0, column=0)
        self.first_name_entry = tk.Entry(user_info_frame)
        self.first_name_entry.grid(row=1, column=0)
        self.first_name_entry.bind("<Return>", Utility.focus_next_widget)
        
        last_name_label = tk.Label(user_info_frame, text="Last Name (required)")
        last_name_label.grid(row=0, column=1)
        self.last_name_entry = tk.Entry(user_info_frame)
        self.last_name_entry.grid(row=1, column=1)
        self.last_name_entry.bind("<Return>", Utility.focus_next_widget)

        title_label = tk.Label(user_info_frame, text="Title")
        title_label.grid(row=0, column=2)
        self.title_combobox = ttk.Combobox(user_info_frame, values=["Mr.", "Ms.", "Dr."], state="readonly")
        self.title_combobox.set('')
        self.title_combobox.grid(row=1, column=2)
        self.title_combobox.bind("<Return>", Utility.focus_next_widget)

        age_label = tk.Label(user_info_frame, text="Age")
        age_label.grid(row=2, column=0)
        self.age_combobox = ttk.Combobox(user_info_frame, values=[str(age) for age in range(15, 101)], state="readonly")
        self.age_combobox.grid(row=3, column=0)
        self.age_combobox.bind("<Return>", Utility.focus_next_widget)

        nationality_label = tk.Label(user_info_frame, text="Nationality")
        nationality_label.grid(row=2, column=1)
        self.nationality_combobox = ttk.Combobox(user_info_frame, values=["French", "Tunisian", "Brazilian", "Ukrainian", "Polish", "Swiss", "Lebanese", "Other"], state="readonly")
        self.nationality_combobox.set('')
        self.nationality_combobox.grid(row=3, column=1)
        self.nationality_combobox.bind("<Return>", Utility.focus_next_widget)

        for widget in user_info_frame.winfo_children():
            widget.grid_configure(padx=10, pady=5)

        # Runner Info Section
        runner_info_frame = tk.LabelFrame(self, text="Runner Information")
        runner_info_frame.pack(padx=20, pady=10, fill="x")

        self.runner_type_var = tk.StringVar(value="")
        runner_type_label = tk.Label(runner_info_frame, text="Type of Runner (required)")
        runner_type_label.grid(row=0, column=0)
        student_button = tk.Radiobutton(runner_info_frame, text="Student", variable=self.runner_type_var, value="Student")
        student_button.grid(row=1, column=0)
        professor_button = tk.Radiobutton(runner_info_frame, text="Professor", variable=self.runner_type_var, value="Professor")
        professor_button.grid(row=1, column=1)
        student_button.bind("<Return>", Utility.focus_next_widget) # those can't be focused with the tab key
        professor_button.bind("<Return>", Utility.focus_next_widget) # those can't be modified by pressing a key

        school_label = tk.Label(runner_info_frame, text="School (required)")
        school_label.grid(row=0, column=2)
        self.school_combobox = ttk.Combobox(runner_info_frame, values=["Ensta Paris", "Télécom Paris", "ENSAE", "Polytechnique", "AgroParistech", "Centrale Supelec", "ENS Paris-Saclay", "Other"], state="readonly")
        self.school_combobox.grid(row=1, column=2)
        self.school_combobox.bind("<Return>", Utility.focus_next_widget)

        for widget in runner_info_frame.winfo_children():
            widget.grid_configure(padx=10, pady=5)

        # Username and Password Info
        login_info_frame = tk.LabelFrame(self, text="Login Information")
        login_info_frame.pack(padx=20, pady=10, fill="x")

        username_label = tk.Label(login_info_frame, text="Username (required)")
        username_label.grid(row=0, column=0)
        self.username_entry = tk.Entry(login_info_frame)
        self.username_entry.grid(row=1, column=0)
        self.username_entry.bind("<Return>", Utility.focus_next_widget)

        password_label = tk.Label(login_info_frame, text="Password (required)")
        password_label.grid(row=0, column=1)
        self.password_entry = tk.Entry(login_info_frame, show="*")
        self.password_entry.grid(row=1, column=1)
        self.password_entry.bind("<Return>", Utility.focus_next_widget)

        password_confirm_label = tk.Label(login_info_frame, text="Confirm Password (required)")
        password_confirm_label.grid(row=0, column=2)
        self.password_confirm_entry = tk.Entry(login_info_frame, show="*")
        self.password_confirm_entry.grid(row=1, column=2)
        self.password_confirm_entry.bind("<Return>", Utility.focus_next_widget)

        for widget in login_info_frame.winfo_children():
            widget.grid_configure(padx=10, pady=5)

        # Accept Terms and Conditions
        terms_frame = tk.LabelFrame(self, text="Terms & Conditions (required)")
        terms_frame.pack(padx=20, pady=10, fill="x")

        self.accept_var = tk.StringVar(value="Not Accepted")
        terms_check = tk.Checkbutton(terms_frame, text="I accept the terms and conditions.", variable=self.accept_var, onvalue="Accepted", offvalue="Not Accepted")
        terms_check.grid(row=0, column=0)
        terms_check.bind("<Return>", Utility.focus_next_widget)

        # Buttons
        self.submit_button = tk.Button(self, text="Submit", command=self.sign_up)
        self.submit_button.pack()
        self.submit_button.bind("<Return>", lambda e: self.submit_button.invoke())

        self.back_button = tk.Button(self, text="Back to Login", command=lambda: self.navigate_callback(1))
        self.back_button.pack()
        self.back_button.bind("<Tab>", lambda e: self.refocus_to_firstname())
        self.back_button.bind("<Return>", lambda e: self.back_button.invoke())

    # Function to sign up a new user and save the data in the Excel file
    def sign_up(self):
        # Check acceptance of terms and conditions
        accepted = self.accept_var.get()
        if accepted == "Accepted":

            # User info
            firstname = self.first_name_entry.get()
            lastname = self.last_name_entry.get()
            # raising error if first name or last name is empty
            if not firstname:
                Utility.show_dismissable_messagebox(self, "Error", "First name is required", lambda: None)
                return
            if not lastname:
                Utility.show_dismissable_messagebox(self, "Error", "Last name is required", lambda: None)
                return
            
            # Runner info (title, age, nationality type of runner, school)
            title = self.title_combobox.get()
            age = self.age_combobox.get()
            nationality = self.nationality_combobox.get()
            # --> those are optional
            type_of_runner = self.runner_type_var.get()
            school = self.school_combobox.get()
            # raising error if type of runner, school, nationality is empty
            if type_of_runner == "":
                Utility.show_dismissable_messagebox(self, "Error", "Type of runner is required", lambda: None)
                return
            if school == "":
                Utility.show_dismissable_messagebox(self, "Error", "School is required", lambda: None)
                return

            # Username and password
            username = self.username_entry.get()
            password = self.password_entry.get()
            password_confirm = self.password_confirm_entry.get()
            # raising error if self.username or password is empty
            if not username or not password or not password_confirm:
                Utility.show_dismissable_messagebox(self, "Error", "Username and passwords are required", lambda: None)
                return
            # raising error if password and confirm password do not match
            if password != password_confirm:
                Utility.show_dismissable_messagebox(self, "Error", "Passwords do not match", lambda: None)
                return
            # the password should be at least 6 characters long, contain 2 numbers, and include at least one special character from *_+=?&@#-
            if not self.validate_password(password):
                Utility.show_dismissable_messagebox(self, "Error", "Password must be at least 6 characters long, contain 2 numbers, and include at least one special character from *_+=?&@#-", lambda: None)
                return

            # Filling the data in the excel file
            filepath = "runner_data.xlsx"
            
            # Check if file exists, otherwise create a new one
            if not os.path.exists(filepath):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                heading = ["Title", "First Name", "Last Name", "Age", "Nationality", "Type of Runner", "School", "Username", "Password"]
                sheet.append(heading)
                workbook.save(filepath)

            # Load the workbook and sheet
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            
            # Check if the data already exists in the excel file (Full duplicate check)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if (row[1] == firstname and row[2] == lastname) or row[7] == username:
                    Utility.show_dismissable_messagebox(self, "Error", "This user already exists", lambda: None)
                    return
                if row[8] == password:  # Check for duplicate password
                    Utility.show_dismissable_messagebox(self, "Error", "This password is already used", lambda: None)
                    return

            # If no duplicates found, add new entry
            sheet.append([title, firstname, lastname, age, nationality, type_of_runner, school, username, password])
            workbook.save(filepath)

            # Reset the fields after submission
            self.first_name_entry.delete(0, 'end')
            self.last_name_entry.delete(0, 'end')
            self.title_combobox.set("")
            self.age_combobox.set('')
            self.nationality_combobox.set("")
            self.runner_type_var.set("")
            self.school_combobox.set("")
            self.username_entry.delete(0, 'end')
            self.password_entry.delete(0, 'end')
            self.password_confirm_entry.delete(0, 'end')
            self.accept_var.set("Not Accepted")

            Utility.show_dismissable_messagebox(self, "Success", "Sign-up successful", lambda: self.navigate_callback(1))
            self.navigate_callback(1)
        
        # If the terms are not accepted, show a warning
        else:
            Utility.show_dismissable_messagebox(self, "Error", "You have not accepted the terms", lambda: None)
    
      # Function to focus the next widget in the tab order when the Enter key is pressed
  
    # Function to validate the password
    def validate_password(self, password):
        return len(password) >= 8 and len(re.findall(r'[0-9]', password)) >= 2 and len(re.findall(r'[a-z]', password))+len(re.findall(r'[A-Z]', password)) >= 6 and any(char in "*_+=?&@#-" for char in password)

    # Function to refocus on the first name entry field when the Tab key is pressed on the Sign Up button
    def refocus_to_firstname(self):
        self.first_name_entry.focus()
        return "break"  



# Screen 3: Blank Screen
class BlankScreen(tk.Frame):
    def __init__(self, master, navigate_callback):
        super().__init__(master)
        self.navigate_callback = navigate_callback

        # Header Section
        self.header_frame = tk.Frame(self, bg="#FFDB58", height=100, width = window.winfo_width())
        self.header_frame.pack(side="top", fill="x")
        self.header_frame.pack_propagate(False)

        self.logo = PhotoImage(file="App_Icon.png").subsample(20, 20)
        # Home Button
        self.home_button = tk.Button(
            self.header_frame,
            text="My App",
            image= self.logo,
            compound="left",
            font=("Helvetica", 18, "bold"),
            bg="#FFDB58",
            command=lambda: self.navigate_callback(1),  # Navigate to the home screen
        )
        self.home_button.pack(side="left", padx=10)

        # Event Presentation Section
        self.event_frame = tk.Frame(self, bg="#5cfaff")
        self.event_frame.pack(fill="x", pady=10)
        self.event_text = tk.Label(
            self.event_frame,
            text="Welcome to the Annual Relay Race!\nExperience the thrill and excitement with teams competing in this challenging course.",
            font=("Helvetica", 14),
            wraplength=800,
            bg="#5cfaff",
            justify="center",
        )
        self.event_text.pack(pady=10)
        self.event_image = PhotoImage(file="5k.png").subsample(2, 2)
        self.event_image_label = tk.Label(self.event_frame, image=self.event_image, bg="#5cfaff")
        self.event_image_label.pack()

        # Search Bar
        self.search_var = tk.StringVar()
        self.search_bar = tk.Entry(
            self.header_frame,
            textvariable=self.search_var,
            font=("Helvetica", 14),
            highlightthickness=2,
            highlightcolor="black",
        )
        self.search_bar.pack(side="right", padx=10, pady=10)
        self.search_bar.bind("<Return>", self.perform_search)
        self.scroll_frame_canvas = tk.Canvas(window)
        self.scroll_frame_canvas.pack(side="top", fill="both", expand=True)
        self.scrollbar = tk.Scrollbar(self.scroll_frame_canvas, orient="vertical", command=self.scroll_frame_canvas.yview)
        self.scrollbar.pack(side="right", fill="y")
        self.scroll_frame_canvas.config(yscrollcommand=self.scrollbar.set)
        self.scrollable_frame = tk.Frame(self.scroll_frame_canvas, bg="#5cfaff", highlightthickness=0, takefocus=False)
        self.scroll_frame_canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.scroll_frame_canvas.bind_all("<MouseWheel>", self.on_mouse_wheel)

    def on_mouse_wheel(self, event):
        """Function for mouse wheel scrolling"""
        if event.delta:
            self.scroll_frame_canvas.yview_scroll(-1 if event.delta > 0 else 1, "units")

    def perform_search(self, event=None):
        search_query = self.search_var.get()

        # Clear the previous results
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()

        teams = Teams("Ekiden_resultats.txt")
        distance_relay_runners = Teams.Distance_Relay(search_query, 7.2, Teams.team_type.MIXED, teams.get_teams())
        for runner in distance_relay_runners.distance_relay:
            runner_info = (
                f"Name: {runner.name}, Sex: {runner.sex.name}, "
                f"Passage Number: {runner.passage_number}, Distance: {runner.distance} km, "
                f"Time: {runner.medley_time} seconds, Speed: {runner.speed:.2f} km/h"
            )
            runner_label = tk.Label(self.scrollable_frame, text=runner_info, wraplength=500, anchor="w", justify="left")
            runner_label.pack(pady=2, anchor="w")

        # Update the scrollable area after adding the widgets
        self.scrollable_frame.update_idletasks()
        self.scroll_frame_canvas.config(scrollregion=self.scroll_frame_canvas.bbox("all"))

       

# Main Scrollable App
class ScrollableApp:
    def __init__(self, window):
        self.window = window
        self.window.title("Multi-Screen Scrollable App")

        # Set the window to full screen
        self.window.attributes("-fullscreen", True)
        window.configure(height=window.winfo_height(), width=window.winfo_width())
        Icone_app = PhotoImage(file="App_Icon.png").subsample(20, 20)
        self.window.iconphoto(True, Icone_app)

        # Create a container frame to hold all screens
        self.container = tk.Frame(window)
        self.container.pack(fill="both", expand=True)

        # Create and store screens in the container
        self.screens = {
            1: LoginScreen(self.container, self.show_screen),
            2: SignupScreen(self.container, self.show_screen),
            3: BlankScreen(self.container, self.show_screen)
        }

        # Place all screens in the same location within the container
        for screen in self.screens.values():
            screen.grid(row=0, column=0, sticky="nsew")

        # Start with the first screen
        self.current_screen = 1
        self.show_screen(self.current_screen)

        # Bind keyboard inputs
        self.window.bind("<Escape>", self.quit_game)

    # Function to show the selected screen
    def show_screen(self, screen_number):
        # Raise the selected screen to the front and update the current screen tracker
        self.screens[screen_number].tkraise()
        self.current_screen = screen_number

    # Function to navigate to the next screen
    def next_screen(self, event=None):
        if self.current_screen < len(self.screens):
            self.show_screen(self.current_screen + 1)

    # Function to navigate to the previous screen
    def previous_screen(self, event=None):
        if self.current_screen > 1:
            self.show_screen(self.current_screen - 1)

    # Function to quit the app
    def quit_game(self, event=None):
        # Quit the app
        self.window.quit()



# Create the main window and lauch the app
window = tk.Tk()
app = ScrollableApp(window)
window.mainloop()
