import tkinter as tk
from tkinter import ttk, messagebox
import os
import openpyxl
import re



# Utility class to gather some useful functions that are common to multiple screens
class Utility:
    @staticmethod
    # Function to show a dismissable messagebox with a custom duration, that can navigate to a new screen also, depending on the callback
    def show_dismissable_messagebox(parent, title, message, navigate_callback, duration = 3000):
        current_focus = parent.focus_get()

        # Create a top-level window
        popup = tk.Toplevel(parent)
        popup.title(title)
        popup.geometry("300x150")  
        popup.transient(parent)  
        popup.resizable(False, False)  
        popup.grab_set()

        # Add the message to the popup
        label = tk.Label(popup, text=message, font=("Helvetica", 12), wraplength=250, pady=20)
        label.pack()

        # Add a dismiss button
        dismiss_button = tk.Button(popup, text="Dismiss", command=lambda: (popup.destroy(), navigate_callback()))
        dismiss_button.pack(pady=10)

        # Automatically close the popup after `duration` milliseconds and navigate
        def auto_close():
            if popup.winfo_exists():  # Check if the popup is still open
                popup.destroy()
                navigate_callback()

        popup.after(duration, auto_close)

        # Return focus to the previously focused widget when the popup closes
        def on_close():
            if current_focus:
                current_focus.focus_set()

        popup.protocol("WM_DELETE_WINDOW", lambda: (on_close(), popup.destroy()))  # Handle the close button (X)
        popup.bind("<Destroy>", lambda event: on_close())
    
    @staticmethod
    # Function to focus the next widget in the tab order when the Enter key is pressed
    def focus_next_widget(event):
        # Focus the next widget in tab order
        event.widget.tk_focusNext().focus()
        return "break"



# Screen 1: Login Screen
class LoginScreen(tk.Frame):
    # Initialize the Login Screen
    def __init__(self, master, navigate_callback):
        super().__init__(master)
        self.navigate_callback = navigate_callback

        label = tk.Label(self, text="Login", font=("Helvetica", 24))
        label.pack()

        # Creating the screen fields
        username_label = tk.Label(self, text="Username:")
        username_label.pack()
        self.username_entry = tk.Entry(self)
        self.username_entry.pack()
        self.username_entry.bind("<Return>", Utility.focus_next_widget)

        password_label = tk.Label(self, text="Password:")
        password_label.pack()
        self.password_entry = tk.Entry(self, show="*")
        self.password_entry.pack()
        self.password_entry.bind("<Return>", Utility.focus_next_widget)

        self.login_button = tk.Button(self, text="Login", command=self.login)
        self.login_button.pack()
        self.login_button.bind("<Return>", lambda e: self.login_button.invoke())

        self.signup_button = tk.Button(self, text="Sign Up", command=lambda: self.navigate_callback(2))
        self.signup_button.pack()
        self.signup_button.bind("<Return>", lambda e: self.signup_button.invoke())
        self.signup_button.bind("<Tab>", lambda e: self.refocus_to_username())

        # Focus on the username entry field by default so the user can start typing immediately
        self.after(10, self.refocus_to_username)

    # Function to check if the login credentials are valid (in the Excel file)
    def login(self):
        username = self.username_entry.get()
        password = self.password_entry.get()

        # Open the Excel file and get the active sheet
        try:
            workbook = openpyxl.load_workbook('runner_data.xlsx')  # Adjust the path to your actual Excel file
            sheet = workbook.active
            
            # Loop through each row in the sheet (starting from row 2 to skip the header row)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Check if the input username and password match the stored values
                if row[7] == username and row[8] == password:
                    Utility.show_dismissable_messagebox(self, "Success", "Login successful!", lambda: self.navigate_callback(3))
                    return
            # If no match is found, show an error and reset the password fields
            Utility.show_dismissable_messagebox(self, "Error", "Invalid login", lambda: None)
            self.username_entry.delete(0, 'end')
            self.password_entry.delete(0, 'end')
        
        except Exception as e:
            messagebox.showerror(title="Error", message=f"An error occurred: {e}")

    # Function to refocus on the username entry field when the Tab key is pressed on the Sign Up button
    def refocus_to_username(self):
        self.username_entry.focus()
        return "break"  



# Screen 2: Sign Up Screen
class SignupScreen(tk.Frame):
    # Initialize the Sign Up Screen
    def __init__(self, master, navigate_callback):
        super().__init__(master)
        self.navigate_callback = navigate_callback

        # Title Label
        label = tk.Label(self, text="Sign Up Form", font=("Helvetica", 24))
        label.pack(pady=20)

        # User Info Section
        user_info_frame = tk.LabelFrame(self, text="User Information")
        user_info_frame.pack(padx=20, pady=10, fill="x")

        first_name_label = tk.Label(user_info_frame, text="First Name (required)")
        first_name_label.grid(row=0, column=0)
        self.first_name_entry = tk.Entry(user_info_frame)
        self.first_name_entry.grid(row=1, column=0)
        self.first_name_entry.bind("<Return>", Utility.focus_next_widget)
        
        last_name_label = tk.Label(user_info_frame, text="Last Name (required)")
        last_name_label.grid(row=0, column=1)
        self.last_name_entry = tk.Entry(user_info_frame)
        self.last_name_entry.grid(row=1, column=1)
        self.last_name_entry.bind("<Return>", Utility.focus_next_widget)

        title_label = tk.Label(user_info_frame, text="Title")
        title_label.grid(row=0, column=2)
        self.title_combobox = ttk.Combobox(user_info_frame, values=["Mr.", "Ms.", "Dr."], state="readonly")
        self.title_combobox.set('')
        self.title_combobox.grid(row=1, column=2)
        self.title_combobox.bind("<Return>", Utility.focus_next_widget)

        age_label = tk.Label(user_info_frame, text="Age")
        age_label.grid(row=2, column=0)
        self.age_combobox = ttk.Combobox(user_info_frame, values=[str(age) for age in range(15, 101)], state="readonly")
        self.age_combobox.grid(row=3, column=0)
        self.age_combobox.bind("<Return>", Utility.focus_next_widget)

        nationality_label = tk.Label(user_info_frame, text="Nationality")
        nationality_label.grid(row=2, column=1)
        self.nationality_combobox = ttk.Combobox(user_info_frame, values=["French", "Tunisian", "Brazilian", "Ukrainian", "Polish", "Swiss", "Lebanese", "Other"], state="readonly")
        self.nationality_combobox.set('')
        self.nationality_combobox.grid(row=3, column=1)
        self.nationality_combobox.bind("<Return>", Utility.focus_next_widget)

        for widget in user_info_frame.winfo_children():
            widget.grid_configure(padx=10, pady=5)

        # Runner Info Section
        runner_info_frame = tk.LabelFrame(self, text="Runner Information")
        runner_info_frame.pack(padx=20, pady=10, fill="x")

        self.runner_type_var = tk.StringVar(value="")
        runner_type_label = tk.Label(runner_info_frame, text="Type of Runner (required)")
        runner_type_label.grid(row=0, column=0)
        student_button = tk.Radiobutton(runner_info_frame, text="Student", variable=self.runner_type_var, value="Student")
        student_button.grid(row=1, column=0)
        professor_button = tk.Radiobutton(runner_info_frame, text="Professor", variable=self.runner_type_var, value="Professor")
        professor_button.grid(row=1, column=1)
        student_button.bind("<Return>", Utility.focus_next_widget) # those can't be focused with the tab key
        professor_button.bind("<Return>", Utility.focus_next_widget) # those can't be modified by pressing a key

        school_label = tk.Label(runner_info_frame, text="School (required)")
        school_label.grid(row=0, column=2)
        self.school_combobox = ttk.Combobox(runner_info_frame, values=["Ensta Paris", "Télécom Paris", "ENSAE", "Polytechnique", "AgroParistech", "Centrale Supelec", "ENS Paris-Saclay", "Other"], state="readonly")
        self.school_combobox.grid(row=1, column=2)
        self.school_combobox.bind("<Return>", Utility.focus_next_widget)

        for widget in runner_info_frame.winfo_children():
            widget.grid_configure(padx=10, pady=5)

        # Username and Password Info
        login_info_frame = tk.LabelFrame(self, text="Login Information")
        login_info_frame.pack(padx=20, pady=10, fill="x")

        username_label = tk.Label(login_info_frame, text="Username (required)")
        username_label.grid(row=0, column=0)
        self.username_entry = tk.Entry(login_info_frame)
        self.username_entry.grid(row=1, column=0)
        self.username_entry.bind("<Return>", Utility.focus_next_widget)

        password_label = tk.Label(login_info_frame, text="Password (required)")
        password_label.grid(row=0, column=1)
        self.password_entry = tk.Entry(login_info_frame, show="*")
        self.password_entry.grid(row=1, column=1)
        self.password_entry.bind("<Return>", Utility.focus_next_widget)

        password_confirm_label = tk.Label(login_info_frame, text="Confirm Password (required)")
        password_confirm_label.grid(row=0, column=2)
        self.password_confirm_entry = tk.Entry(login_info_frame, show="*")
        self.password_confirm_entry.grid(row=1, column=2)
        self.password_confirm_entry.bind("<Return>", Utility.focus_next_widget)

        for widget in login_info_frame.winfo_children():
            widget.grid_configure(padx=10, pady=5)

        # Accept Terms and Conditions
        terms_frame = tk.LabelFrame(self, text="Terms & Conditions (required)")
        terms_frame.pack(padx=20, pady=10, fill="x")

        self.accept_var = tk.StringVar(value="Not Accepted")
        terms_check = tk.Checkbutton(terms_frame, text="I accept the terms and conditions.", variable=self.accept_var, onvalue="Accepted", offvalue="Not Accepted")
        terms_check.grid(row=0, column=0)
        terms_check.bind("<Return>", Utility.focus_next_widget)

        # Buttons
        self.submit_button = tk.Button(self, text="Submit", command=self.sign_up)
        self.submit_button.pack()
        self.submit_button.bind("<Return>", lambda e: self.submit_button.invoke())

        self.back_button = tk.Button(self, text="Back to Login", command=lambda: self.navigate_callback(1))
        self.back_button.pack()
        self.back_button.bind("<Tab>", lambda e: self.refocus_to_firstname())
        self.back_button.bind("<Return>", lambda e: self.back_button.invoke())

        # Focus on the username entry field by default so the user can start typing immediately
        self.after(10, self.refocus_to_firstname)

    # Function to sign up a new user and save the data in the Excel file
    def sign_up(self):
        # Check acceptance of terms and conditions
        accepted = self.accept_var.get()
        if accepted == "Accepted":

            # User info
            firstname = self.first_name_entry.get()
            lastname = self.last_name_entry.get()
            # raising error if first name or last name is empty
            if not firstname:
                Utility.show_dismissable_messagebox(self, "Error", "First name is required", lambda: None)
                return
            if not lastname:
                Utility.show_dismissable_messagebox(self, "Error", "Last name is required", lambda: None)
                return
            
            # Runner info (title, age, nationality type of runner, school)
            title = self.title_combobox.get()
            age = self.age_combobox.get()
            nationality = self.nationality_combobox.get()
            # --> those are optional
            type_of_runner = self.runner_type_var.get()
            school = self.school_combobox.get()
            # raising error if type of runner, school, nationality is empty
            if type_of_runner == "":
                Utility.show_dismissable_messagebox(self, "Error", "Type of runner is required", lambda: None)
                return
            if school == "":
                Utility.show_dismissable_messagebox(self, "Error", "School is required", lambda: None)
                return

            # Username and password
            username = self.username_entry.get()
            password = self.password_entry.get()
            password_confirm = self.password_confirm_entry.get()
            # raising error if self.username or password is empty
            if not username or not password or not password_confirm:
                Utility.show_dismissable_messagebox(self, "Error", "Username and passwords are required", lambda: None)
                return
            # raising error if password and confirm password do not match
            if password != password_confirm:
                Utility.show_dismissable_messagebox(self, "Error", "Passwords do not match", lambda: None)
                return
            # the password should be at least 6 characters long, contain 2 numbers, and include at least one special character from *_+=?&@#-
            if not self.validate_password(password):
                Utility.show_dismissable_messagebox(self, "Error", "Password must be at least 6 characters long, contain 2 numbers, and include at least one special character from *_+=?&@#-", lambda: None)
                return

            # Filling the data in the excel file
            filepath = "runner_data.xlsx"
            
            # Check if file exists, otherwise create a new one
            if not os.path.exists(filepath):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                heading = ["Title", "First Name", "Last Name", "Age", "Nationality", "Type of Runner", "School", "Username", "Password"]
                sheet.append(heading)
                workbook.save(filepath)

            # Load the workbook and sheet
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            
            # Check if the data already exists in the excel file (Full duplicate check)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if (row[1] == firstname and row[2] == lastname) or row[7] == username:
                    Utility.show_dismissable_messagebox(self, "Error", "This user already exists", lambda: None)
                    return
                if row[8] == password:  # Check for duplicate password
                    Utility.show_dismissable_messagebox(self, "Error", "This password is already used", lambda: None)
                    return

            # If no duplicates found, add new entry
            sheet.append([title, firstname, lastname, age, nationality, type_of_runner, school, username, password])
            workbook.save(filepath)

            # Reset the fields after submission
            self.first_name_entry.delete(0, 'end')
            self.last_name_entry.delete(0, 'end')
            self.title_combobox.set("")
            self.age_combobox.set('')
            self.nationality_combobox.set("")
            self.runner_type_var.set("")
            self.school_combobox.set("")
            self.username_entry.delete(0, 'end')
            self.password_entry.delete(0, 'end')
            self.password_confirm_entry.delete(0, 'end')
            self.accept_var.set("Not Accepted")

            Utility.show_dismissable_messagebox(self, "Success", "Sign-up successful", lambda: self.navigate_callback(1))
            self.navigate_callback(1)
        
        # If the terms are not accepted, show a warning
        else:
            Utility.show_dismissable_messagebox(self, "Error", "You have not accepted the terms", lambda: None)
    
      # Function to focus the next widget in the tab order when the Enter key is pressed
  
    # Function to validate the password
    def validate_password(self, password):
        return len(password) >= 8 and len(re.findall(r'[0-9]', password)) >= 2 and len(re.findall(r'[a-z]', password))+len(re.findall(r'[A-Z]', password)) >= 6 and any(char in "*_+=?&@#-" for char in password)

    # Function to refocus on the first name entry field when the Tab key is pressed on the Sign Up button
    def refocus_to_firstname(self):
        self.first_name_entry.focus()
        return "break"  



# Screen 3: Blank Screen
class BlankScreen(tk.Frame):
    def __init__(self, master, navigate_callback):
        super().__init__(master)
        self.navigate_callback = navigate_callback

        self.label = tk.Label(self, text="Welcome to Screen 3!", font=("Helvetica", 24))
        self.label.pack(pady=20)



# Main Scrollable App
class ScrollableApp:
    def __init__(self, window):
        self.window = window
        self.window.title("Multi-Screen Scrollable App")

        # Set the window to full screen
        self.window.attributes("-fullscreen", True)

        # Create a container frame to hold all screens
        self.container = tk.Frame(window)
        self.container.pack(fill="both", expand=True)

        # Create and store screens in the container
        self.screens = {
            1: LoginScreen(self.container, self.show_screen),
            2: SignupScreen(self.container, self.show_screen),
            3: BlankScreen(self.container, self.show_screen)
        }

        # Place all screens in the same location within the container
        for screen in self.screens.values():
            screen.grid(row=0, column=0, sticky="nsew")

        # Start with the first screen
        self.current_screen = 1
        self.show_screen(self.current_screen)

        # Bind keyboard inputs
        self.window.bind("<Escape>", self.quit_game)

    # Function to show the selected screen
    def show_screen(self, screen_number):
        # Raise the selected screen to the front and update the current screen tracker
        self.screens[screen_number].tkraise()
        self.current_screen = screen_number

    # Function to navigate to the next screen
    def next_screen(self, event=None):
        if self.current_screen < len(self.screens):
            self.show_screen(self.current_screen + 1)

    # Function to navigate to the previous screen
    def previous_screen(self, event=None):
        if self.current_screen > 1:
            self.show_screen(self.current_screen - 1)

    # Function to quit the app
    def quit_game(self, event=None):
        # Quit the app
        self.window.quit()



# Create the main window and lauch the app
window = tk.Tk()
app = ScrollableApp(window)
window.mainloop()
