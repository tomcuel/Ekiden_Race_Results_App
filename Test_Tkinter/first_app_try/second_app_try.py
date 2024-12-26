from enum import Enum
import tkinter as tk
from tkinter import ttk, messagebox
from tkinter import PhotoImage, Scrollbar, Canvas
import os
import openpyxl
import re



# Define the teams_data class to parse the data and store the team information
class Teams:

    # Enum to define the team categories
    class team_type(Enum):
        MEN = 0
        WOMEN = 1
        MIXED = 2 # Admin is included in the mixed category
        UNKOWN = 4


    # Enum to define the fields of the team_info object
    class field(Enum):
        TEAM_NAME = 0
        TEAM_CATEGORY = 1
        TEAM_TIME = 2
        TEAM_RANKING_IN_CATEGORY = 3
        TEAM_OVERALL_RANKING = 4


    # Define the runner class to store the data of each runner
    class runner:
        def __init__(self, name: str, sex: int, passage_number: int, distance: float, medley_time: int, speed: float):
            self.name = name
            self.sex = Teams.team_type(sex)
            self.passage_number = passage_number
            self.distance = distance
            self.medley_time = medley_time  # Time in seconds
            self.speed = speed

    # Define the team_info object to store the data of each team
    class team_info:
        # Initialize the team_info object with the team's name, category, time, runners, and ranking
        def __init__(self, name: str, category: int, time: int, category_ranking: int, ranking: int, runners: list):
            self.data = {
                Teams.field.TEAM_NAME: name,
                Teams.field.TEAM_CATEGORY: category,
                Teams.field.TEAM_TIME: time,
                Teams.field.TEAM_RANKING_IN_CATEGORY: category_ranking,
                Teams.field.TEAM_OVERALL_RANKING: ranking
            }
            self.runners = runners 


    # Initialize the teams_data object with the file path
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.teams = self._parse_teams()

    # Parse the teams data from the file, return a list of team_info objects
    def _parse_teams(self):
        with open(self.file_path, "r") as file:
            data = file.read().strip().split("\n")

        # Remove the first 5 lines of unnecessary data
        data = data[5:]
        number_of_teams = data.count("") + 1

        # Initialize lists to store each team's data
        teams_data_parsing = [[] for _ in range(number_of_teams)]
        team_index = 0
        while data:
            line = data.pop(0)
            if line == "":
                team_index += 1
                continue
            teams_data_parsing[team_index].append(line)

        # Process each team into team_info objects
        teams = []
        for i, team_info_parsing in enumerate(teams_data_parsing):
            if not team_info_parsing:
                continue
            
            # The first line gives us the data of the team
            team_info_line = team_info_parsing[0].strip().split("|")
            
            # Extract the team name
            team_name = team_info_line[0]
            
            # Get the ekiden total time in seconds
            team_time_str = team_info_line[1].strip().split(" ")[3].split(":")
            team_time = int(team_time_str[0])*3600 + int(team_time_str[1])*60 + int(team_time_str[2])
            
            # Get the team category number
            # Non classed teams have an extra word in the category
            delta = 0
            if team_info_line[2].strip().split(" ")[0] == "Non" :
                delta = 1
            team_category_str = team_info_line[2].strip().split(" ")[1+delta]
            if team_category_str == "Hommes":
                category = Teams.team_type.MEN
            elif team_category_str == "Dames":
                category = Teams.team_type.WOMEN
            elif team_category_str == "Mixte":
                category = Teams.team_type.MIXED
            elif team_category_str == "Admin":
                category = Teams.team_type.MIXED
            else:
                category = Teams.team_type.UNKOWN
            
            # Get the team ranking in its category
            category_team_ranking = team_info_line[2].strip().split(" ")[0][:2]
            if category_team_ranking[1] == '°':
                category_team_ranking = category_team_ranking[0]
            if category_team_ranking == "No":
                category_team_ranking = -1
            else :
                category_team_ranking = int(category_team_ranking)
            
            # Get the team overall ranking
            team_ranking = i+1

            runners = []
            # Process the runners data, 6 runners for line 3/4/5/6/7/8
            for i in range(2, 8) : 
                
                runner_info = team_info_parsing[i].split("\t")

                # Name and Sex
                runner_name = runner_info[1].strip()
                runner_sex = Teams.team_type.MEN if runner_info[2] == "M" else Teams.team_type.WOMEN

                # Runner passage number
                runner_passage_number = i-1

                 # Runner distance
                runner_distance = 7.2
                if (runner_passage_number == 1 or runner_passage_number == 3 or runner_passage_number == 5):
                    runner_distance = 5.0
                elif (runner_passage_number == 2 or runner_passage_number == 4):
                    runner_distance = 10.0

                # Medley time 
                medley_time_str = runner_info[4]
                medley_time = 0
                # Check if the time includes hours (separating the hours, minutes, and seconds differently if there is hours or not)
                if not 'X' in medley_time_str:
                    if 'h' in medley_time_str:
                        medley_time = 3600 * int(medley_time_str[0]) + 60 * int(medley_time_str[2:4]) + int(medley_time_str[5:7])
                    else:
                        medley_time = 60 * int(medley_time_str[0:2]) + int(medley_time_str[3:5])
            
                # Speed
                runner_speed_str = runner_info[5].strip()
                if runner_speed_str and runner_speed_str.replace(",", ".").replace(".", "", 1).isdigit():
                    runner_speed = float(runner_speed_str.replace(",", "."))
                else:
                    runner_speed = 0.0  # Or handle as needed (e.g., raise an error or set a default value)

                # Append the runner to the runners list
                partial_runner = Teams.runner(name=runner_name, sex=runner_sex, passage_number=runner_passage_number, distance=runner_distance, medley_time=medley_time, speed=runner_speed)
                runners.append(partial_runner)


            # Create a team_info instance and add it to the teams list
            team = Teams.team_info(name=team_name, category=category, time=team_time, category_ranking=category_team_ranking, ranking=team_ranking, runners=runners)
            teams.append(team)

        # Return the list of team_info objects
        return teams

    # Function to get the teams
    def get_teams(self):
        return self.teams
    
    # Function to get the team based on its name (or at least a part of it)
    def get_team(self, team_name: str):
        for team in self.teams:
            if team_name in team.data[Teams.field.TEAM_NAME]:
                return team
        return None
    
    # Function to get the teams of one category
    def get_teams_by_category(self, category):
        return [team for team in self.teams if team.data[Teams.field.TEAM_CATEGORY] == category]
    
    # Function to get the teams of a category depending on the team name
    def get_teams_by_category_by_name(self, name : str):
        category_searched = -1
        for team in self.teams : 
            if name in team.data[Teams.field.TEAM_NAME] :
                category_searched = team.data[Teams.field.TEAM_CATEGORY]
                break
        return self.get_teams_by_category(category_searched)
    
    # Function to get the runners of a team
    def get_runners(self, team):
        return team.runners
    

    # Define the number relay class to store the data of each relay
    class Number_Relay:

        # Initialize the number relay object with the number of the relay and teams data
        def __init__(self, name : str, number: int, sex: int, teams: list):
            self.name = name
            self.number = number
            self.sex = sex
            self.teams = teams
            self.number_relay = self.get_number_relay(name, number, sex)

        # function to get the distance relay
        # the sex matter, because it will get the scratch time of the runners, or the ordering based on the sex
        def get_number_relay(self, name : str, number : int,  sex : int):
            self.name = name    
            self.number = number
            self.sex = sex
            relay_number = []
            # empty name means we want to get the relay based on the number, still looking at the sex
            if name == "":
                for team in self.teams:
                    if team.runners[self.number - 1].sex == sex or sex == Teams.team_type.MIXED : 
                        relay_number.append(team.runners[self.number - 1])
            # otherwise we want to get the relay based on the name, still looking at the sex
            else : 
                number_searched = -1
                sex_searched = -1
                for team in self.teams : 
                    for runner in team.runners : 
                        if name in runner.name : 
                            number_searched = runner.passage_number
                            self.number = number_searched
                            sex_searched = runner.sex
                            self.sex = sex_searched
                            break
                # if the number is not found, return an empty list since there is no runner with this name
                if number_searched == -1 :
                    return []
                for team in self.teams:
                    if team.runners[number_searched - 1].sex == sex or sex == Teams.team_type.MIXED : 
                        relay_number.append(team.runners[number_searched - 1])
            # return the list of the runners
            return relay_number

    # Define the distance relay class to store the data of each relay
    class Distance_Relay:

        # Initialize the distance relay object with the distance and teams data
        def __init__(self, name : str, distance: float, sex: int, teams: list):
            self.name = name
            self.distance = distance
            self.sex = sex
            self.teams = teams
            self.distance_relay = self.get_distance_relay(name, distance, sex)

        # function to get the distance relay
        # the sex matter, because it will get the scratch time of the runners, or the ordering based on the sex
        # no name means that we want to get the relay based on the distance, still looking at the sex
        # if we want a WOMEN but the name correspond to a MEN, it will return the MIXED relay
        def get_distance_relay(self, name : str, distance : float, sex : int):
            self.name = name
            self.distance = distance
            self.sex = sex
            relay_distance = []
            # empty name means we want to get the relay based on the distance, still looking at the sex
            if name == "":
                for team in self.teams:
                    for runner in team.runners : 
                        if runner.distance == distance and (runner.sex == sex or sex == Teams.team_type.MIXED) : 
                            relay_distance.append(runner)
            # otherwise we want to get the relay based on the name, still looking at the sex
            else : 
                distance_searched = -1
                sex_searched = -1
                for team in self.teams : 
                    for runner in team.runners : 
                        if name in runner.name :
                            distance_searched = runner.distance
                            self.distance = distance_searched
                            sex_searched = runner.sex
                            self.sex = sex_searched
                            break
                # if the distance is not found, return an empty list since there is no runner with this name
                if distance_searched == -1 or sex_searched == -1 :
                    return []
                for team in self.teams:
                    for runner in team.runners : 
                        if runner.distance == distance_searched and (runner.sex == sex_searched or sex == Teams.team_type.MIXED) : 
                            relay_distance.append(runner)
            # return the list of the runners
            return relay_distance



# Utility class to gather some useful functions that are common to multiple screens
class Utility:
    @staticmethod
    # Function to show a dismissable messagebox with a custom duration, that can navigate to a new screen also, depending on the callback
    def show_dismissable_messagebox(parent, title, message, navigate_callback, duration = 3000, is_deconnexion_avorted = False):
        current_focus = parent.focus_get()
        
        popup = tk.Toplevel(parent)
        popup.title(title)
        window_width = 500
        window_height = 250
        screen_width = parent.winfo_screenwidth()
        screen_height = parent.winfo_screenheight()
        position_top = int(screen_height / 2 - window_height / 2)
        position_right = int(screen_width / 2 - window_width / 2)
        popup.geometry(f'{window_width}x{window_height}+{position_right}+{position_top}')
        popup.transient(parent) 
        popup.resizable(False, False)  
        popup.grab_set()

        # Add text to the popup
        canvas = tk.Canvas(popup, width=window_width, height=window_height / 2, bg="#282c34", highlightthickness=0)
        canvas.pack(fill="both", expand=True)
        canvas.create_text(
            window_width // 2, window_height // 4, text=message, font=("Helvetica", 40), fill="#ffffff", width=window_width - 40, anchor="center"
        )

        # Add dismiss button
        dismiss_button = tk.Button(
            popup, text="Ok", font=("Helvetica", 30), fg="#000000", relief="flat",  activeforeground="#ff0000", border=0, highlightthickness=2.5, 
            command=lambda: (popup.destroy(), navigate_callback())
        )
        dismiss_button.place(
            relx=0.5, rely=0.75, anchor="center", width=200, height=50
        )

        # Automatically close the popup after `duration` milliseconds (optional)
        def auto_close():
            if popup.winfo_exists():
                popup.destroy()
                if not is_deconnexion_avorted : # if the deconnexion is not aborted, we can navigate to the next screen, otherwise we stay on the same screen (the 3rd one)
                    navigate_callback()
        
        # Close the popup after `duration` milliseconds
        popup.after(duration, auto_close)
            
        # Return focus to the previously focused widget when the popup closes
        def on_close():
            if current_focus:
                current_focus.focus_set()
    
        # Handle the close button
        popup.protocol("WM_DELETE_WINDOW", lambda: (on_close(), popup.destroy()))  # Handle the close button (X)
        popup.bind("<Destroy>", lambda event: on_close())
    
    @staticmethod
    # Function to focus the next widget in the tab order when the Enter key is pressed
    def focus_next_widget(event):
        # Focus the next widget in tab order
        event.widget.tk_focusNext().focus()
        return "break"



# Screen 1: Login Screen
class LoginScreen(tk.Frame):
    # Initialize the Login Screen
    def __init__(self, master, navigate_callback):
        super().__init__(master, bg="#282c34", highlightthickness=0, borderwidth=0, border=0)
        self.navigate_callback = navigate_callback
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # Header Label
        label = tk.Label(
            self, text="Ekiden Analysis App", font=("Helvetica", 100, "bold"), fg="#ffffff", bg="#282c34"
        )
        label.pack(pady=(90, 50))

        # Username Field
        username_label = tk.Label(
            self, text="Username:", font=("Helvetica", 40), fg="#ffffff", bg="#282c34"
        )
        username_label.pack()
        self.username_entry = tk.Entry(
            self, font=("Helvetica", 20), bg="#f7f7f7"
        )
        self.username_entry.pack(pady=(30, 20))
        self.username_entry.bind("<Return>", Utility.focus_next_widget)

        # Password Field
        password_label = tk.Label(
            self, text="Password:", font=("Helvetica", 40), fg="#ffffff", bg="#282c34"
        )
        password_label.pack()
        self.password_entry = tk.Entry(
            self, show="*", font=("Helvetica", 20), bg="#f7f7f7"
        )
        self.password_entry.pack(pady=(30, 20))
        self.password_entry.bind("<Return>", Utility.focus_next_widget)

        # Login Button 
        self.login_button = tk.Button(
            self, text="Login", font=("Helvetica", 40), fg="#000000", relief="flat", activeforeground="#ff0000", border=0, highlightthickness=2.5, command=self.login,
        )
        self.login_button.pack(pady=(20, 20))
        self.login_button.bind("<Return>", lambda e: self.login_button.invoke())

        # Sign Up Button
        self.signup_button = tk.Button(
            self, text="Sign Up", font=("Helvetica", 40), fg="#000000", relief="flat", activeforeground="#ff0000", border=0, highlightthickness=2.5, command=lambda: [self.navigate_callback(2), self.clear_user_info()] 
        )
        self.signup_button.pack(pady=(20, 20))
        self.signup_button.bind("<Return>", lambda e: self.signup_button.invoke())
        self.signup_button.bind("<Tab>", lambda e: self.refocus_to_username())

    # Function to check if the login credentials are valid (in the Excel file)
    def login(self):
        username = self.username_entry.get()
        password = self.password_entry.get()

        # Open the Excel file and get the active sheet
        try:
            workbook = openpyxl.load_workbook('runner_data.xlsx')  # Adjust the path to your actual Excel file
            sheet = workbook.active
            
            # Loop through each row in the sheet (starting from row 2 to skip the header row)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Check if the input username and password match the stored values
                if row[7] == username and row[8] == password:
                    Utility.show_dismissable_messagebox(self, "Success", "Login successful!", lambda: self.navigate_callback(3))
                    self.clear_user_info()
                    return
            # If no match is found, show an error and reset the password fields
            Utility.show_dismissable_messagebox(self, "Error", "Invalid login", lambda: None)
            self.clear_user_info()
        
        except Exception as e:
            messagebox.showerror(title="Error", message=f"An error occurred: {e}")

    # Function to refocus on the username entry field when the Tab key is pressed on the Sign Up button
    def refocus_to_username(self):
        self.username_entry.focus()
        return "break"
    
    # Function to clear both username and password fields from the login screen
    def clear_user_info(self):
        self.username_entry.delete(0, 'end')
        self.password_entry.delete(0, 'end')



# Screen 2: Sign Up Screen
class SignupScreen(tk.Frame):
    # Initialize the Sign Up Screen
    def __init__(self, master, navigate_callback):
        super().__init__(master, bg="#282c34", highlightthickness=0, borderwidth=0, border=0)
        self.navigate_callback = navigate_callback
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # Title Label
        label = tk.Label(
            self, text="Sign Up", font=("Helvetica", 70, "bold"), fg="#ffffff", bg="#282c34"
            )
        label.pack(pady=(20, 20))   

        # User Info Section
        user_info_frame = tk.LabelFrame(
            self, text="User Information", font=("Helvetica", 30), fg="#ffffff", bg="#282c34", border=0, labelanchor="n"
        )
        user_info_frame.pack(pady=(10,30), fill="y")

        first_name_label = tk.Label(
            user_info_frame, text="First Name (required)", font=("Helvetica", 20), fg="#ffffff", bg="#282c34"
        )
        first_name_label.grid(row=0, column=0)
        self.first_name_entry = tk.Entry(user_info_frame)
        self.first_name_entry.grid(row=1, column=0)
        self.first_name_entry.bind("<Return>", Utility.focus_next_widget)
        
        last_name_label = tk.Label(
            user_info_frame, text="Last Name (required)", font=("Helvetica", 20), fg="#ffffff", bg="#282c34"
        )
        last_name_label.grid(row=0, column=1)
        self.last_name_entry = tk.Entry(user_info_frame)
        self.last_name_entry.grid(row=1, column=1)
        self.last_name_entry.bind("<Return>", Utility.focus_next_widget)

        title_label = tk.Label(
            user_info_frame, text="Title", font=("Helvetica", 20), fg="#ffffff", bg="#282c34"
        )
        title_label.grid(row=0, column=2)
        self.title_combobox = ttk.Combobox(
            user_info_frame, values=["Mr.", "Ms.", "Dr."], state="readonly"
        )
        self.title_combobox.set('')
        self.title_combobox.grid(row=1, column=2)
        self.title_combobox.bind("<Return>", Utility.focus_next_widget)

        age_label = tk.Label(
            user_info_frame, text="Age", font=("Helvetica", 20), fg="#ffffff", bg="#282c34"
        )
        age_label.grid(row=2, column=0)
        self.age_combobox = ttk.Combobox(
            user_info_frame, values=[str(age) for age in range(15, 101)], state="readonly"
        )
        self.age_combobox.grid(row=3, column=0)
        self.age_combobox.bind("<Return>", Utility.focus_next_widget)

        nationality_label = tk.Label(
            user_info_frame, text="Nationality", font=("Helvetica", 20), fg="#ffffff", bg="#282c34"
            )
        nationality_label.grid(row=2, column=1)
        self.nationality_combobox = ttk.Combobox(
            user_info_frame, values=["French", "Tunisian", "Brazilian", "Ukrainian", "Polish", "Swiss", "Lebanese", "Other"], state="readonly"
            )
        self.nationality_combobox.set('')
        self.nationality_combobox.grid(row=3, column=1)
        self.nationality_combobox.bind("<Return>", Utility.focus_next_widget)

        for widget in user_info_frame.winfo_children():
            widget.grid_configure(padx=10, pady=5)

        # Runner Info Section
        runner_info_frame = tk.LabelFrame(
            self, text="Runner Information (required)", font=("Helvetica", 30), fg="#ffffff", bg="#282c34", border=0, labelanchor="n"
        )
        runner_info_frame.pack(pady=(10,30), fill="y")

        school_label = tk.Label(
            runner_info_frame, text="School", font=("Helvetica", 20), fg="#ffffff", bg="#282c34"
        )
        school_label.grid(row=0, column=0)
        self.school_combobox = ttk.Combobox(
            runner_info_frame, values=["Ensta Paris", "Télécom Paris", "ENSAE", "Polytechnique", "AgroParistech", "Centrale Supelec", "ENS Paris-Saclay", "Other"], state="readonly"
        )
        self.school_combobox.grid(row=1, column=0)
        self.school_combobox.bind("<Return>", Utility.focus_next_widget)

        self.runner_type_var = tk.StringVar(value="")
        runner_type_label = tk.Label(
            runner_info_frame, text="Type of Runner ", font=("Helvetica", 20), fg="#ffffff", bg="#282c34"
        )
        runner_type_label.grid(row=0, column=1, columnspan=2)
        student_button = tk.Radiobutton(
            runner_info_frame, text="Student", font=("Helvetica", 20), fg="#ffffff", bg="#282c34", variable=self.runner_type_var, value="Student"
        )
        student_button.grid(row=1, column=1)
        professor_button = tk.Radiobutton(
            runner_info_frame, text="Professor", font=("Helvetica", 20), fg="#ffffff", bg="#282c34", variable=self.runner_type_var, value="Professor"
        )
        professor_button.grid(row=1, column=2)
        student_button.bind("<Return>", Utility.focus_next_widget) # those can't be focused with the tab key
        professor_button.bind("<Return>", Utility.focus_next_widget) # those can't be modified by pressing a key

        for widget in runner_info_frame.winfo_children():
            widget.grid_configure(padx=10, pady=5)

        # Username and Password Info
        login_info_frame = tk.LabelFrame(
            self, text="Login Information (required)", font=("Helvetica", 30), fg="#ffffff", bg="#282c34", border=0, labelanchor="n"
        )
        login_info_frame.pack(pady=(10,30), fill="y")

        username_label = tk.Label(
            login_info_frame, text="Username", font=("Helvetica", 20), fg="#ffffff", bg="#282c34"
        )
        username_label.grid(row=0, column=0)
        self.username_entry = tk.Entry(login_info_frame)
        self.username_entry.grid(row=1, column=0)
        self.username_entry.bind("<Return>", Utility.focus_next_widget)

        password_label = tk.Label(
            login_info_frame, text="Password", font=("Helvetica", 20), fg="#ffffff", bg="#282c34"
        )
        password_label.grid(row=0, column=1)
        self.password_entry = tk.Entry(login_info_frame, show="*")
        self.password_entry.grid(row=1, column=1)
        self.password_entry.bind("<Return>", Utility.focus_next_widget)

        password_confirm_label = tk.Label(
            login_info_frame, text="Confirm Password", font=("Helvetica", 20), fg="#ffffff", bg="#282c34"
        )
        password_confirm_label.grid(row=0, column=2)
        self.password_confirm_entry = tk.Entry(login_info_frame, show="*")
        self.password_confirm_entry.grid(row=1, column=2)
        self.password_confirm_entry.bind("<Return>", Utility.focus_next_widget)

        for widget in login_info_frame.winfo_children():
            widget.grid_configure(padx=10, pady=5)

        # Accept Terms and Conditions
        terms_frame = tk.LabelFrame(
            self, text="Terms & Conditions (required)", font=("Helvetica", 20), fg="#ffffff", bg="#282c34"
        )
        terms_frame.pack(padx=20, pady=10, fill="x")
        self.accept_var = tk.StringVar(value="Not Accepted")
        terms_check = tk.Checkbutton(
            terms_frame, text="I accept the terms and conditions", font=("Helvetica", 10), fg="#ffffff", bg="#282c34",
            variable=self.accept_var, onvalue="Accepted", offvalue="Not Accepted"
        )
        terms_check.grid(row=0, column=0)
        terms_check.bind("<Return>", Utility.focus_next_widget)

        # Buttons
        self.submit_button = tk.Button(
            self, text="Submit", font=("Helvetica", 30), fg="#000000", relief="flat", activeforeground="#ff0000", border=0, highlightthickness=2.5, command=self.sign_up
        )
        self.submit_button.pack(pady=(20, 20))
        self.submit_button.bind("<Return>", lambda e: self.submit_button.invoke())

        self.back_button = tk.Button(
            self, text="Back to Login", font=("Helvetica", 30), fg="#000000", relief="flat", activeforeground="#ff0000", border=0, highlightthickness=2.5, command=lambda: self.navigate_callback(1),
        )
        self.back_button.pack(pady=(20, 20))
        self.back_button.bind("<Return>", lambda e: self.back_button.invoke())
        self.back_button.bind("<Tab>", lambda e: self.refocus_to_firstname())

    # Function to sign up a new user and save the data in the Excel file
    def sign_up(self):
        # Check acceptance of terms and conditions
        accepted = self.accept_var.get()
        if accepted == "Accepted":

            # User info
            firstname = self.first_name_entry.get()
            lastname = self.last_name_entry.get()
            # raising error if first name or last name is empty
            if not firstname:
                Utility.show_dismissable_messagebox(self, "Error", "First name is required", lambda: None)
                return
            if not lastname:
                Utility.show_dismissable_messagebox(self, "Error", "Last name is required", lambda: None)
                return
            
            # Runner info (title, age, nationality type of runner, school)
            title = self.title_combobox.get()
            age = self.age_combobox.get()
            nationality = self.nationality_combobox.get()
            # --> those are optional
            type_of_runner = self.runner_type_var.get()
            school = self.school_combobox.get()
            # raising error if type of runner, school, nationality is empty
            if type_of_runner == "":
                Utility.show_dismissable_messagebox(self, "Error", "Type of runner is required", lambda: None)
                return
            if school == "":
                Utility.show_dismissable_messagebox(self, "Error", "School is required", lambda: None)
                return

            # Username and password
            username = self.username_entry.get()
            password = self.password_entry.get()
            password_confirm = self.password_confirm_entry.get()
            # raising error if self.username or password is empty
            if not username or not password or not password_confirm:
                Utility.show_dismissable_messagebox(self, "Error", "Username and passwords are required", lambda: None)
                return
            # raising error if password and confirm password do not match
            if password != password_confirm:
                Utility.show_dismissable_messagebox(self, "Error", "Passwords do not match", lambda: None)
                return
            # the password should be at least 6 characters long, contain 2 numbers, and include at least one special character from *_+=?&@#-
            if not self.validate_password(password):
                Utility.show_dismissable_messagebox(self, "Error", "Password must be at least 6 characters long, contain 2 numbers, and include at least one special character from *_+=?&@#-", lambda: None)
                return

            # Filling the data in the excel file
            filepath = "runner_data.xlsx"
            
            # Check if file exists, otherwise create a new one
            if not os.path.exists(filepath):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                heading = ["Title", "First Name", "Last Name", "Age", "Nationality", "Type of Runner", "School", "Username", "Password"]
                sheet.append(heading)
                workbook.save(filepath)

            # Load the workbook and sheet
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            
            # Check if the data already exists in the excel file (Full duplicate check)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if (row[1] == firstname and row[2] == lastname) or row[7] == username:
                    Utility.show_dismissable_messagebox(self, "Error", "This user already exists", lambda: None)
                    return
                if row[8] == password:  # Check for duplicate password
                    Utility.show_dismissable_messagebox(self, "Error", "This password is already used", lambda: None)
                    return

            # If no duplicates found, add new entry
            sheet.append([title, firstname, lastname, age, nationality, type_of_runner, school, username, password])
            workbook.save(filepath)

            # Reset the fields after submission
            self.first_name_entry.delete(0, 'end')
            self.last_name_entry.delete(0, 'end')
            self.title_combobox.set("")
            self.age_combobox.set('')
            self.nationality_combobox.set("")
            self.runner_type_var.set("")
            self.school_combobox.set("")
            self.username_entry.delete(0, 'end')
            self.password_entry.delete(0, 'end')
            self.password_confirm_entry.delete(0, 'end')
            self.accept_var.set("Not Accepted")

            Utility.show_dismissable_messagebox(self, "Success", "Sign-up successful", lambda: self.navigate_callback(1))
            self.navigate_callback(1)
        
        # If the terms are not accepted, show a warning
        else:
            Utility.show_dismissable_messagebox(self, "Error", "You have not accepted the terms", lambda: None)
    
      # Function to focus the next widget in the tab order when the Enter key is pressed
  
    # Function to validate the password
    def validate_password(self, password):
        return len(password) >= 8 and len(re.findall(r'[0-9]', password)) >= 2 and len(re.findall(r'[a-z]', password))+len(re.findall(r'[A-Z]', password)) >= 6 and any(char in "*_+=?&@#-" for char in password)

    # Function to refocus on the first name entry field when the Tab key is pressed on the Sign Up button
    def refocus_to_firstname(self):
        self.first_name_entry.focus()
        return "break"  



# Screen 3: Main App Screen
class Main_App_Screen(tk.Frame):
    def __init__(self, master, navigate_callback, window_width):
        super().__init__(master, bg="#282c34", highlightthickness=0, borderwidth=0, border=0)
        self.navigate_callback = navigate_callback
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

# Header Section
        header_frame = tk.Frame(self, bg="#FFDB58", height=100, width = window_width)
        header_frame.pack(side="top", fill="x")
        header_frame.pack_propagate(True)

    # Home Icon Button
        self.logo = PhotoImage(file="App_Icon.png").subsample(20, 20)
        logo_button = tk.Button(
            header_frame, image= self.logo, compound="left", border=0, highlightthickness=0
        )
        logo_button.place(x=100, y=header_frame.winfo_reqheight() // 2, anchor="center")  
        logo_button.bind("<Button-1>", self.reset_screen)

    # Home Label/Button
        home_button = tk.Label(
            header_frame, text="Home", font=("Helvetica", 40), bg="#FFDB58", border=0, highlightthickness=0
        )
        home_button.place(x=300, y=header_frame.winfo_reqheight() // 2, anchor="center") 
        home_button.bind("<Enter>", lambda event: home_button.config(font=("Helvetica", 50, "bold"), fg="#000000", bg="#FFDB58"))
        home_button.bind("<Leave>", lambda event: home_button.config(font=("Helvetica", 40), fg="#000000", bg="#FFDB58"))
        home_button.bind("<Button-1>", self.reset_screen)
        
    # Results Label/Button
        results_button = tk.Label(
            header_frame, text="Results", font=("Helvetica", 40), bg="#FFDB58", border=0, highlightthickness=0
        )
        results_button.place(x=600, y=header_frame.winfo_reqheight() // 2, anchor="center")
        results_button.bind("<Button-1>", lambda event: print("Results_showing!"))

    # Result Sub-Menu Buttons
        # utility functions
        def show_submenu(event=None):
            # Display the submenu when hovering over the results button.
            submenu.place(x=results_button.winfo_x(), y=results_button.winfo_y() + results_button.winfo_height())
            submenu.lift()
            for sub_button in submenu_buttons:
                sub_button.update_idletasks() # Ensure the text on buttons is refreshed immediately
            results_button.config(font=("Helvetica", 50, "bold")) # Result button bigger when submenu is showing
        def hide_submenu(event=None):
            # Hide the submenu if the mouse is not over the submenu or results button.
            widget_under_pointer = master.winfo_containing(event.x_root, event.y_root)
            if widget_under_pointer not in submenu.winfo_children() and widget_under_pointer != submenu and widget_under_pointer != results_button:
                submenu.place_forget()
            results_button.config(font=("Helvetica", 40)) # Result button smaller when submenu isn't showing
        def on_submenu_button_hover(event, button, is_entering):
            # Change visuals when hovering over submenu buttons.
            if is_entering:
                button.config(font=("Helvetica", 30, "bold"))
            else:
                button.config(font=("Helvetica", 20))

        # Submenu Frame with fixed size
        submenu = tk.Frame(master, bg="#FFDB58", relief="raised", border=0, highlightthickness=0, width=300, height=100)
        submenu.place_forget()

        # Add submenu buttons with place
        team_result_sub_button = tk.Label(
                submenu, text="Team Results",  font=("Helvetica", 30),  bg="#FFDB58", fg="#000000", padx=10, pady=5, anchor="center"
            )
        team_result_sub_button.place(x=0, y=0, relwidth=1, height=50)
        team_result_sub_button.bind("<Enter>", lambda event, b=team_result_sub_button: on_submenu_button_hover(event, b, True))
        team_result_sub_button.bind("<Leave>", lambda event, b=team_result_sub_button: on_submenu_button_hover(event, b, False))
        team_result_sub_button.bind("<Button-1>", lambda event: print(" Teams_results_showing!"))
        indiv_result_sub_button = tk.Label(
                submenu, text="Individual Results", font=("Helvetica", 30), bg="#FFDB58", fg="#000000", padx=10, pady=5, anchor="center"
            )
        indiv_result_sub_button.place(x=0, y=50, relwidth=1, height=50)
        indiv_result_sub_button.bind("<Enter>", lambda event, b=indiv_result_sub_button: on_submenu_button_hover(event, b, True))
        indiv_result_sub_button.bind("<Leave>", lambda event, b=indiv_result_sub_button: on_submenu_button_hover(event, b, False))
        indiv_result_sub_button.bind("<Button-1>", lambda event: print("Individuals_results_showing!"))
        submenu_buttons = [team_result_sub_button, indiv_result_sub_button]

    # Bind events to the results and submenu buttons
        results_button.bind("<Enter>", show_submenu)
        results_button.bind("<Leave>", hide_submenu)
        submenu.bind("<Enter>", show_submenu)
        submenu.bind("<Leave>", hide_submenu)

    # Deconnexion Button
        # function to use the dismissable messagebox to ask the user if he really wants to deconnect
        def deconnexion(event=None):
            Utility.show_dismissable_messagebox(master, "Deconnexion", "Are you sure you want to deconnect?", lambda: self.navigate_callback(1), duration=4000, is_deconnexion_avorted=True)
        
        self.logout = PhotoImage(file="Log_Out.png").subsample(20, 20)
        logout_button = tk.Button(
            header_frame, image=self.logout, border=0, highlightthickness=0, compound="center"
        )
        logout_button.place(x=window_width - 100, y=header_frame.winfo_reqheight() // 2, anchor="center")
        logout_button.bind("<Button-1>", deconnexion)

    # Search Bar
        self.search_var = tk.StringVar()
        search_bar = tk.Entry(
            header_frame, textvariable=self.search_var, font=("Helvetica", 20), background="#f7f7f7", highlightbackground="#282c34"
        )
        search_bar.place(x=window_width // 2 + 275, y=header_frame.winfo_reqheight() // 2, anchor="center")
        search_bar.bind("<Return>", self.perform_search)

        # Create the Scrollable Frame for Results
        self.scroll_canvas = tk.Canvas(self, bg="#282c34", highlightthickness=0)
        self.scroll_canvas.pack(side="left", fill="both", expand=True)
        scrollbar = tk.Scrollbar(self, orient="vertical", command=self.scroll_canvas.yview)
        scrollbar.pack(side="right", fill="y")
        self.scroll_canvas.config(yscrollcommand=scrollbar.set)

        # Mouse Wheel Scrolling
        self.scroll_canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        
        # Frame for fixed header content (hidden initially)
        self.results_header_frame = tk.Frame(self.scroll_canvas, bg="#282c34", height=40)
        self.results_header_window = self.scroll_canvas.create_window(
            (0, 0), window=self.results_header_frame, anchor="nw", width=self.scroll_canvas.winfo_width()
        )

        # Initial hidden header for the results
        self.results_header_frame.pack_forget()
        header_labels = ["Rank", "Name", "Category", "Time", "Speed"]
        for index, header_text in enumerate(header_labels):
            tk.Label(
                self.results_header_frame, text=header_text, font=("Helvetica", 20), fg="#ffffff", bg="#282c34", anchor="w", padx=5,
            ).grid(row=0, column=index, sticky="nsew")
        for col in range(len(header_labels)):
            self.results_header_frame.grid_columnconfigure(col, weight=1)

        # Frame for scrollable results
        self.scrollable_frame = tk.Frame(self.scroll_canvas, bg="#282c34")
        self.scrollable_frame.bind(
            "<Configure>", lambda e: self.scroll_canvas.config(scrollregion=self.scroll_canvas.bbox("all"))
        )
        self.scroll_canvas.create_window((0, 40), window=self.scrollable_frame, anchor="nw")

        # Placeholder Text
        self.placeholder_label = tk.Label(
            self.scrollable_frame, text="Welcome! Please enter a query in the search bar.",font=("Helvetica", 20), bg="#282c34", fg="#FFFFFF"
        )
        self.placeholder_label.pack()

    # function to make the search and display the results in the scrollable frame
    def perform_search(self, event=None):
        search_query = self.search_var.get().strip()

        # Clear previous results or placeholder text
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()

        if search_query:
            self.placeholder_label.pack_forget()

            # Show the header when results are found
            self.results_header_frame.pack(fill="x", side="top")

            teams = Teams("Ekiden_resultats.txt")
            distance_relay_runners = Teams.Distance_Relay(search_query, 7.2, Teams.team_type.MIXED, teams.get_teams())

            # Display results in a tabular format
            for idx, runner in enumerate(distance_relay_runners.distance_relay, start=1):
                runner_info = [
                    idx,
                    runner.name,
                    runner.sex.name,
                    f"{runner.medley_time:.2f} s",
                    f"{runner.speed:.2f} km/h",
                ]
                for col, text in enumerate(runner_info):
                    tk.Label(
                        self.scrollable_frame,
                        text=text,
                        font=("Helvetica", 12),
                        bg="#f7f7f7" if idx % 2 == 0 else "#ffffff",
                        anchor="w",
                        padx=5,
                        pady=2,
                    ).grid(row=idx, column=col, sticky="nsew")

                # Configure columns to expand
                for col in range(len(runner_info)):
                    self.scrollable_frame.grid_columnconfigure(col, weight=1)
        else:
            # Reset to default state
            self.placeholder_label.pack(expand=True)
            self.results_header_frame.pack_forget()  # Hide the header when no results

    # handle mouse wheel scrolling
    def _on_mousewheel(self, event):
        self.scroll_canvas.yview_scroll(-1 if event.delta > 0 else 1, "units")

    # reset the scollable frame to its initial state (the text and gpx preview)
    def reset_screen(self, event=None):
        # Clear the search bar
        self.search_var.set("")

        # Clear previous results or placeholder text
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()

        # Hide the results header
        self.results_header_frame.pack_forget()

        # Ensure placeholder_label is reset
        self.placeholder_label_new = tk.Label(
                self.scrollable_frame,
                text="Welcome! Please enter a query in the search bar.",
                font=("Helvetica", 20),
                bg="#282c34",
                fg="#FFFFFF",
        )
        
        # Use place() to position the label
        self.placeholder_label_new.pack()

        # Optionally, scroll back to the top of the canvas
        self.scroll_canvas.yview_moveto(0)
       




# Main Scrollable App
class ScrollableApp:
    def __init__(self, window):
        self.window = window
        self.window.title("Multi-Screen Scrollable App")
        self.window.attributes("-fullscreen", True)
        window.configure(height=window.winfo_height(), width=window.winfo_width())
        Icone_app = PhotoImage(file="App_Icon.png")
        self.window.iconphoto(True, Icone_app)

        # Create a container frame to hold all screens
        self.container = tk.Frame(window)
        self.container.pack(fill="both", expand=True)

        # Create and store screens in the container
        self.screens = {
            1: LoginScreen(self.container, self.show_screen),
            2: SignupScreen(self.container, self.show_screen),
            3: Main_App_Screen(self.container, self.show_screen, window.winfo_width())
        }

        # Place all screens in the same location within the container
        for screen in self.screens.values():
            screen.grid(row=0, column=0, sticky="nsew")

        # Start with the first screen
        self.current_screen = 3 
        ''' 1 '''
        self.show_screen(self.current_screen)

        # Bind keyboard inputs
        self.window.bind("<Escape>", self.quit_game)

    # Function to show the selected screen
    def show_screen(self, screen_number):
        # Raise the selected screen to the front and update the current screen tracker
        self.screens[screen_number].tkraise()
        self.current_screen = screen_number
        # direct focus to the first field for the login and sign up screens
        if self.current_screen == 1:
            self.screens[screen_number].after(500, lambda: self.screens[screen_number].username_entry.focus())
        elif self.current_screen == 2:
            self.screens[screen_number].after(500, lambda: self.screens[screen_number].first_name_entry.focus())

    # Function to navigate to the next screen
    def next_screen(self, event=None):
        if self.current_screen < len(self.screens):
            self.show_screen(self.current_screen + 1)

    # Function to navigate to the previous screen
    def previous_screen(self, event=None):
        if self.current_screen > 1:
            self.show_screen(self.current_screen - 1)

    # Function to quit the app
    def quit_game(self, event=None):
        # Quit the app
        self.window.quit()



# Create the main window and lauch the app
window = tk.Tk()
app = ScrollableApp(window)
window.mainloop()